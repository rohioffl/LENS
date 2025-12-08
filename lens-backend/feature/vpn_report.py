"""Utility helpers to produce Excel VPN reports."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any, Dict, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _auto_fit(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)


def _serialize(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    try:
        return json.dumps(value, ensure_ascii=False)
    except TypeError:
        return str(value)


def _format_value(value: Any) -> Any:
    """Return a cell-friendly representation with readable newlines for lists."""
    if isinstance(value, list):
        if not value:
            return ""
        str_items = []
        for item in value:
            if isinstance(item, Mapping):
                return _serialize(value)
            str_items.append(_serialize(item))
        return "\n".join(str_items)
    if isinstance(value, tuple):
        return _format_value(list(value))
    return _serialize(value)


def _apply_table_style(ws, header_rows: int = 1) -> None:
    if ws.max_row == 0 or ws.max_column == 0:
        return

    header_font = Font(name="Lato", bold=True, size=12)
    body_font = Font(name="Lato", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    band_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_side = Side(border_style="thin", color="D0D7DE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row_idx in range(1, ws.max_row + 1):
        is_header = row_idx <= header_rows
        use_font = header_font if is_header else body_font
        fill = header_fill if is_header else None
        if not is_header and header_rows:
            if (row_idx - header_rows) % 2 == 1:
                fill = band_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = use_font
            cell.alignment = alignment
            cell.border = border
            if fill:
                cell.fill = fill


def _write_table(ws, headers: Iterable[str], rows: Iterable[Iterable[Any]]) -> None:
    headers = list(headers)
    for header in headers:
        if header is None:
            raise ValueError("Header titles cannot be None")
    if headers:
        ws.append(headers)
    else:
        raise ValueError("At least one header column is required")

    for row in rows:
        ws.append([_format_value(value) for value in row])

    _apply_table_style(ws, header_rows=1)
    _auto_fit(ws)


def build_vpn_report_workbook(context: Dict[str, Any], metadata: Dict[str, Any] | None = None) -> bytes:
    """Return an XLSX workbook (bytes) summarizing VPN inputs/results."""
    aws = context.get("aws", {})
    gcp = context.get("gcp", {})
    vpn = context.get("vpn", {})
    wb = Workbook()

    overview = wb.active
    overview.title = "Overview"
    prefix_value = ""
    if metadata and metadata.get("prefix"):
        prefix_value = metadata["prefix"]
    else:
        prefix_value = vpn.get("name_prefix") or ""

    overview_rows: list[tuple[str, Any]] = [
        ("Run Prefix", prefix_value),
        ("Run Timestamp", metadata.get("timestamp", "") if metadata else ""),
        ("AWS VPC ID", aws.get("id", "")),
        ("AWS VPC Name", aws.get("name", "")),
        ("AWS CIDR", aws.get("cidr", "")),
        ("AWS Region", aws.get("region", "")),
        ("AWS ASN", aws.get("asn") or (metadata or {}).get("config", {}).get("aws_asn", "")),
        ("GCP Project", gcp.get("project", "")),
        ("GCP Network", gcp.get("name", "")),
        ("GCP Region", gcp.get("region", "")),
        ("GCP ASN", gcp.get("asn") or (metadata or {}).get("config", {}).get("gcp_asn", "")),
        ("Placeholder GCP CIDR", vpn.get("placeholder_cidr", "")),
    ]
    _write_table(overview, ["Field", "Value"], overview_rows)

    aws_sheet = wb.create_sheet("AWS Subnets")
    aws_rows = [
        (
            subnet.get("id", ""),
            subnet.get("name", ""),
            subnet.get("cidr", ""),
            subnet.get("az", ""),
        )
        for subnet in aws.get("subnets", [])
    ]
    _write_table(aws_sheet, ["Subnet ID", "Name", "CIDR", "AZ"], aws_rows)

    gcp_sheet = wb.create_sheet("GCP Subnets")
    selected = gcp.get("selected_subnets") or gcp.get("subnetworks", [])
    gcp_rows = [
        (
            subnet.get("name", ""),
            subnet.get("cidr") or subnet.get("ip_cidr_range") or "",
            subnet.get("region", ""),
        )
        for subnet in selected
    ]
    _write_table(gcp_sheet, ["Name", "CIDR", "Region"], gcp_rows)

    vpn_sheet = wb.create_sheet("Resource Names")
    vpn_rows = [
        (key, value)
        for key, value in sorted(vpn.items())
        if not key.endswith("_tf")
    ]
    _write_table(vpn_sheet, ["Identifier", "Value"], vpn_rows)

    if metadata:
        config_sheet = wb.create_sheet("Deployment Config")
        config_rows = [
            (key.replace("_", " ").title(), value)
            for key, value in sorted(metadata.get("config", {}).items())
        ]
        if metadata.get("prefix"):
            config_rows.insert(0, ("Name Prefix", metadata["prefix"]))
        if metadata.get("timestamp"):
            config_rows.insert(1 if metadata.get("prefix") else 0, ("Timestamp", metadata["timestamp"]))
        _write_table(config_sheet, ["Setting", "Value"], config_rows)

        aws_tunnels = metadata.get("aws_tunnels", [])
        if aws_tunnels:
            tunnel_sheet = wb.create_sheet("AWS Tunnels")
            tunnel_rows = [
                (
                    idx,
                    tunnel.get("outside_ip", ""),
                    tunnel.get("customer_inside_ip", ""),
                    tunnel.get("vpn_inside_ip", ""),
                    tunnel.get("bgp_asn", ""),
                    tunnel.get("psk", ""),
                )
                for idx, tunnel in enumerate(aws_tunnels, start=1)
            ]
            _write_table(
                tunnel_sheet,
                [
                    "Tunnel #",
                    "AWS Outside IP",
                    "Customer Inside IP",
                    "AWS Inside IP",
                    "BGP ASN",
                    "IKEv2 PSK",
                ],
                tunnel_rows,
            )

        resources = metadata.get("resources", {})
        gcp_tunnels = resources.get("gcp_tunnels", [])
        if gcp_tunnels:
            gcp_tunnel_sheet = wb.create_sheet("GCP Tunnels")
            gcp_tunnel_rows = [
                (
                    tunnel.get("name", ""),
                    tunnel.get("aws_interface", ""),
                    tunnel.get("gcp_interface", ""),
                    tunnel.get("router_bgp_ip", ""),
                    tunnel.get("peer_bgp_ip", ""),
                    tunnel.get("link", ""),
                )
                for tunnel in gcp_tunnels
            ]
            _write_table(
                gcp_tunnel_sheet,
                [
                    "Tunnel Name",
                    "AWS Interface",
                    "GCP Interface",
                    "Router BGP IP",
                    "Peer BGP IP",
                    "Console Link",
                ],
                gcp_tunnel_rows,
            )

        interface_ips = resources.get("gcp_interface_ips")
        if interface_ips:
            iface_sheet = wb.create_sheet("GCP Interfaces")
            iface_rows = [
                (f"Interface {idx}", ip)
                for idx, ip in enumerate(interface_ips, start=1)
            ]
            _write_table(iface_sheet, ["Interface", "Public IP"], iface_rows)

        aws_resources = []
        gcp_resources = []
        other_resources = []
        for key, value in resources.items():
            if key == "gcp_tunnels" or key == "gcp_interface_ips":
                continue
            entry = (key, _format_value(value))
            if key.startswith("aws_"):
                aws_resources.append(entry)
            elif key.startswith("gcp_"):
                gcp_resources.append(entry)
            else:
                other_resources.append(entry)

        if aws_resources:
            aws_resource_sheet = wb.create_sheet("AWS Resources")
            _write_table(aws_resource_sheet, ["Resource", "Value"], aws_resources)

        if gcp_resources:
            gcp_resource_sheet = wb.create_sheet("GCP Resources")
            _write_table(gcp_resource_sheet, ["Resource", "Value"], gcp_resources)

        if other_resources:
            misc_sheet = wb.create_sheet("Shared Resources")
            _write_table(misc_sheet, ["Resource", "Value"], other_resources)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
