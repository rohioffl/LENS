import argparse
import base64
import concurrent.futures
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from googleapiclient.discovery import build
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================
# 1. CONFIGURATION
# ============================

GCP_SCOPES = [
    'https://www.googleapis.com/auth/cloud-platform',
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets',
]


def build_credentials(service_key):
    if not service_key:
        raise ValueError("GCP service account JSON is required.")
    if isinstance(service_key, str):
        raw = service_key.strip()
        try:
            info = json.loads(raw)
        except json.JSONDecodeError:
            decoded = base64.b64decode(raw).decode("utf-8")
            info = json.loads(decoded)
    elif isinstance(service_key, dict):
        info = service_key
    else:
        raise ValueError("Service key must be JSON or base64-encoded JSON.")
    return service_account.Credentials.from_service_account_info(info, scopes=GCP_SCOPES)


def list_projects(creds):
    crm = build('cloudresourcemanager', 'v1', credentials=creds, cache_discovery=False)
    projects = []
    request = crm.projects().list()
    while request is not None:
        response = request.execute()
        for project in response.get('projects', []):
            if project.get('lifecycleState') != 'ACTIVE':
                continue
            projects.append({
                'project_id': project.get('projectId'),
                'display_name': project.get('name') or project.get('projectId'),
            })
        request = crm.projects().list_next(previous_request=request, previous_response=response)
    return projects


def resolve_project_selection(creds, selected_ids=None):
    if selected_ids:
        return [
            {"project_id": str(pid).strip(), "display_name": str(pid).strip()}
            for pid in selected_ids
            if str(pid).strip()
        ]
    return list_projects(creds)


# ============================
# 2. GOOGLE API HELPERS
# ============================


def get_gcp_services(project_id, creds):
    crm = build('cloudresourcemanager', 'v1', credentials=creds, cache_discovery=False)
    iam = build('iam', 'v1', credentials=creds, cache_discovery=False)
    compute = build('compute', 'v1', credentials=creds, cache_discovery=False)
    storage = build('storage', 'v1', credentials=creds, cache_discovery=False)
    return crm, iam, compute, storage

def get_google_sheet_services(creds):
    sheets_service = build("sheets", "v4", credentials=creds)
    drive_service = build("drive", "v3", credentials=creds)
    return sheets_service, drive_service

# ============================
# 3. AUDIT FUNCTIONS
# ============================
def fetch_iam_users(project_id, crm, client_name):
    user_roles = defaultdict(list)
    deleted_users = []
    try:
        policy = crm.projects().getIamPolicy(resource=project_id, body={}).execute()
        for binding in policy.get("bindings", []):
            clean_role = binding.get("role")
            for member in binding.get("members", []):
                if member.startswith("user:"):
                    email = member.replace("user:", "")
                    user_roles[email].append(clean_role)
                elif member.startswith("deleted:user:"):
                    match = re.match(r'deleted:user:([^?]+)', member)
                    if match:
                        email = match.group(1)
                        deleted_users.append({
                            "Users/Groups": email + " (deleted)",
                            "Name": email.split("@")[0].replace(".", " ").title() + " (deleted)",
                            "Role": f"• {clean_role.split('/')[-1]}",
                            "Comments from Ankercloud": "User deleted in Google Workspace but still has roles attached",
                            f"Comments from {client_name}": ""
                        })
    except Exception as e:
        print(f"❌ IAM Users for {project_id}: {e}")

    records = []
    for email, roles in user_roles.items():
        name = email.split("@")[0].replace(".", " ").title()
        if any(r.lower() in ["roles/editor", "roles/owner"] for r in roles):
            comment = "The user has predefined Editor/Owner role. As a best practice, assign least privilege access."
        else:
            comment = "No action required"
        formatted_roles = "\n".join(f"• {r.split('/')[-1]}" for r in roles) if roles else "-"
        records.append({
            "Users/Groups": email,
            "Name": name,
            "Role": formatted_roles,
            "Comments from Ankercloud": comment,
            f"Comments from {client_name}": ""
        })
    records.extend(deleted_users)
    columns = ["S No", "Users/Groups", "Name", "Role", "Comments from Ankercloud", f"Comments from {client_name}"]
    if not records:
        return pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(records)
        df.index += 1
        df.reset_index(inplace=True)
        df.rename(columns={"index": "S No"}, inplace=True)
        return df

def fetch_iam_groups(project_id, crm, client_name):
    group_roles = defaultdict(list)
    try:
        policy = crm.projects().getIamPolicy(resource=project_id, body={}).execute()
        for binding in policy.get("bindings", []):
            clean_role = binding.get("role")
            for member in binding.get("members", []):
                if member.startswith("group:"):
                    email = member.replace("group:", "")
                    group_roles[email].append(clean_role)
    except Exception as e:
        print(f"❌ IAM Groups for {project_id}: {e}")

    columns = ["S No", "Users/Groups", "Name", "Role", "Comments from Ankercloud", f"Comments from {client_name}"]
    records = []
    for email, roles in group_roles.items():
        name = email.split("@")[0].replace(".", " ").title()
        if any(r.lower() in ["roles/editor", "roles/owner"] for r in roles):
            comment = "The user has predefined Editor/Owner role. As a best practice, assign least privilege access."
        else:
            comment = "No action required"
        formatted_roles = "\n".join(f"• {r.split('/')[-1]}" for r in roles) if roles else "-"
        records.append({
            "Users/Groups": email,
            "Name": name,
            "Role": formatted_roles,
            "Comments from Ankercloud": comment,
            f"Comments from {client_name}": ""
        })
    if not records:
        return pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(records)
        df.index += 1
        df.reset_index(inplace=True)
        df.rename(columns={"index": "S No"}, inplace=True)
        return df

def fetch_service_accounts(project_id, client_name, crm, iam):
    sa_roles = defaultdict(list)
    try:
        policy = crm.projects().getIamPolicy(resource=project_id, body={}).execute()
        for binding in policy.get("bindings", []):
            for member in binding.get("members", []):
                if member.startswith("serviceAccount:"):
                    email = member.replace("serviceAccount:", "")
                    sa_roles[email].append(binding.get("role"))
    except Exception as e:
        print(f"❌ Failed to get IAM policy for {project_id}: {e}")

    sa_data = []
    try:
        request = iam.projects().serviceAccounts().list(name=f"projects/{project_id}")
        while request is not None:
            response = request.execute()
            for sa in response.get("accounts", []):
                email = sa["email"]
                roles = sa_roles.get(email, [])
                try:
                    key_info = iam.projects().serviceAccounts().keys().list(
                        name=f"projects/-/serviceAccounts/{email}", keyTypes="USER_MANAGED"
                    ).execute()
                    keys = key_info.get("keys", [])
                    if not keys:
                        key_desc = "No keys"
                        rotate = False
                    else:
                        key = keys[0]
                        dt = datetime.strptime(key["validAfterTime"], "%Y-%m-%dT%H:%M:%SZ")
                        age = (datetime.now(timezone.utc) - dt.replace(tzinfo=timezone.utc)).days
                        rotate = age >= 90
                        key_desc = dt.strftime("%b %d, %Y") + f" (Key ID: {key['name'].split('/')[-1]})"
                except Exception:
                    key_desc = "Permission Denied"
                    rotate = False
                # Build compound comment
                has_editor_owner = any(r.lower() in ["roles/editor", "roles/owner"] for r in roles)
                if has_editor_owner and rotate:
                    comment = "The user has predefined Editor/Owner role. As a best practice, assign least privilege access. and Need to rotate the service account key"
                elif has_editor_owner:
                    comment = "The user has predefined Editor/Owner role. As a best practice, assign least privilege access."
                elif sa.get("disabled"):
                    comment = "The service account is currently disabled. If no longer needed, we can remove it."
                elif rotate:
                    comment = "Need to rotate the service account key"
                else:
                    comment = "No action required"
                formatted_roles = "\n".join(
                    f"• {r.split('/')[-1].replace('.', ' ').replace('_', ' ').title()}"
                    for r in roles if r and r != "-"
                ) or "-"
                sa_data.append({
                    "Emails": email,
                    "Key creation date": key_desc,
                    "Role": formatted_roles,
                    "Status": "Disabled" if sa.get("disabled") else "Enabled",
                    "Comments from Ankercloud": comment,
                    f"Comments from {client_name}": ""
                })
            request = iam.projects().serviceAccounts().list_next(request, response)
    except Exception as e:
        print(f"❌ Failed fetching Service Accounts for {project_id}: {e}")

    columns = ["S No", "Emails", "Key creation date", "Role", "Status", "Comments from Ankercloud", f"Comments from {client_name}"]
    if not sa_data:
        return pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(sa_data)
        df.index += 1
        df.reset_index(inplace=True)
        df.rename(columns={"index": "S No"}, inplace=True)
        return df


def get_instances_by_tags_and_sas(project_id, compute):
    instance_map = defaultdict(list)
    all_instances = []
    request = compute.instances().aggregatedList(project=project_id)
    while request is not None:
        response = request.execute()
        for _, data in response.get('items', {}).items():
            for instance in data.get("instances", []):
                name = instance["name"]
                all_instances.append(name)
                for tag in instance.get("tags", {}).get("items", []):
                    instance_map[f"tag:{tag}"].append(name)
                for sa in instance.get("serviceAccounts", []):
                    email = sa.get("email")
                    if email:
                        instance_map[f"sa:{email}"].append(name)
        request = compute.instances().aggregatedList_next(previous_request=request, previous_response=response)
    return instance_map, all_instances

def fetch_firewall_rules(project_id, compute, client_name):
    from collections import defaultdict
    import pandas as pd

    def is_ports_80_443_only(rules):
        allowed_ports = set()
        for r in rules:
            proto = r.get("IPProtocol", "")
            ports = r.get("ports", [])
            if proto in ("tcp", "") and ports:
                for p in ports:
                    allowed_ports.add(p)
            if proto in ("tcp", "") and (not ports or ports == ["all"]):
                return False
        return allowed_ports and all(p in ("80", "443") for p in allowed_ports)

    # Build VM mappings for unattached check
    instance_tags = {}
    instance_sas = {}
    instance_networks = {}
    try:
        instance_agg = compute.instances().aggregatedList(project=project_id).execute()
    except Exception as e:
        print(f"Failed to list instances for {project_id}: {e}")
        columns = [
            "S No", "Name", "Type", "Targets", "Filters",
            "Protocols / ports", "Action", "Instances attached",
            "Comments from Ankercloud team", f"Comments from {client_name}"
        ]
        return pd.DataFrame(columns=columns)
    for _, zone in instance_agg.get("items", {}).items():
        for inst in zone.get("instances", []):
            name = inst["name"]
            tags = set(inst.get("tags", {}).get("items", []))
            sas = {sa["email"].lower() for sa in inst.get("serviceAccounts", [])}
            networks = [iface["network"].split("/")[-1] for iface in inst.get("networkInterfaces", [])]
            instance_tags[name] = tags
            instance_sas[name] = sas
            instance_networks[name] = networks

    firewall_rules = []
    try:
        rules_api = compute.firewalls().list(project=project_id)
    except Exception as e:
        print(f"Failed to list firewall rules for {project_id}: {e}")
        rules_api = None
    while rules_api is not None:
        try:
            response = rules_api.execute()
        except Exception as e:
            print(f"Failed to read firewall rules for {project_id}: {e}")
            break
        for rule in response.get("items", []):
            name = rule.get("name")
            direction = rule.get("direction", "INGRESS")
            filters_list = rule.get("sourceRanges", []) if direction == "INGRESS" else rule.get("destinationRanges", [])
            target_tags = set(rule.get("targetTags", []))
            target_sas = set(sa.lower() for sa in rule.get("targetServiceAccounts", []))
            network = rule["network"].split("/")[-1]
            rules = rule.get("allowed", []) if "allowed" in rule else rule.get("denied", [])
            action = "Allow" if "allowed" in rule else "Deny"
            ports = []
            proto_ports = []
            all_ports_flag = False

            for r in rules:
                proto = r.get("IPProtocol", "")
                ports_r = r.get("ports", [])
                if not ports_r or ports_r == ["all"]:
                    proto_ports.append("All")
                    all_ports_flag = True
                else:
                    ports.extend(ports_r)
                    proto_ports.append(f"{proto}:{', '.join(ports_r)}")
            ports_formatted = ", ".join(proto_ports) or "All"

            # Find applicable instances
            applies_to_all = not target_tags and not target_sas
            attached = set()
            for inst, tags in instance_tags.items():
                if network not in instance_networks[inst]:
                    continue
                if applies_to_all or target_tags & tags or target_sas & instance_sas.get(inst, set()):
                    attached.add(inst)
            targets = []
            if applies_to_all: targets.append("ALL")
            if target_tags: targets.append("TAG: " + ", ".join(sorted(target_tags)))
            if target_sas: targets.append("SA: " + ", ".join(sorted(target_sas)))
            targets_str = " | ".join(targets) if targets else "-"

            # Exclude: only 80/443
            if is_ports_80_443_only(rules):
                continue

            is_open_to_world = "0.0.0.0/0" in filters_list

            # Comments
            if not attached:
                if all_ports_flag and is_open_to_world:
                    comment_text = (
                        "All ports are allowed and the firewall rule is open to the internet (0.0.0.0/0). "
                        "The rule is not attached to any instance or service account. "
                        "If it is not required, it can be safely deleted with confirmation. "
                        "As a best practice, restrict the rule to only required ports and trusted IPs."
                    )
                elif all_ports_flag:
                    comment_text = (
                        "All ports are allowed for this firewall rule, and it is not attached to any instance or service account. "
                        "If it is not required, it can be safely deleted with confirmation. "
                        "As a best practice, restrict access to only the required ports."
                    )
                elif is_open_to_world:
                    comment_text = (
                        "This firewall rule is open to the internet (0.0.0.0/0) and is not attached to any instance or service account. "
                        "If it is not required, it can be safely deleted with confirmation. "
                        "As a best practice, restrict access to specific trusted IPs."
                    )
                else:
                    comment_text = (
                        "This firewall rule is not attached to any instance or service account. "
                        "If it is not required, it can be safely deleted with confirmation."
                    )
            else:
                if all_ports_flag and is_open_to_world:
                    comment_text = (
                        "All ports are allowed and the firewall rule is open to the internet (0.0.0.0/0). "
                        "As a best practice, restrict the rule to only required ports and trusted IPs."
                    )
                elif all_ports_flag:
                    comment_text = (
                        "All ports are allowed for this firewall rule. "
                        "As a best practice, restrict access to only the required ports."
                    )
                elif is_open_to_world:
                    comment_text = (
                        "This firewall rule is open to the internet (0.0.0.0/0). "
                        "As a best practice, restrict access to specific trusted IPs."
                    )
                else:
                    comment_text = "No action required."

            instances_attached = "\n".join(f"• {n}" for n in sorted(attached)) if attached else "No instance attached to firewall"

            # Only append if action is required
            if comment_text != "No action required.":
                firewall_rules.append({
                    "S No": len(firewall_rules) + 1,
                    "Name": name,
                    "Type": direction,
                    "Targets": targets_str,
                    "Filters": "\n".join(f"• {ip}" for ip in filters_list) or "-",
                    "Protocols / ports": ports_formatted,
                    "Action": action,
                    "Instances attached": instances_attached,
                    "Comments from Ankercloud team": comment_text,
                    f"Comments from {client_name}": ""
                })
        rules_api = compute.firewalls().list_next(previous_request=rules_api, previous_response=response)

    columns = [
        "S No", "Name", "Type", "Targets", "Filters",
        "Protocols / ports", "Action", "Instances attached",
        "Comments from Ankercloud team", f"Comments from {client_name}"
    ]
    if not firewall_rules:
        return pd.DataFrame(columns=columns)
    else:
        return pd.DataFrame(firewall_rules, columns=columns)

def fetch_public_buckets(project_id, storage, client_name):
    bucket_data = []
    try:
        buckets = storage.buckets().list(project=project_id).execute().get('items', [])
        for bucket in buckets:
            try:
                bucket_name = bucket["name"]
                bucket_location = bucket.get("location", "")
                storage_class = bucket.get("storageClass", "")
                try:
                    policy = storage.buckets().getIamPolicy(bucket=bucket_name, userProject=project_id).execute()
                except Exception as e:
                    print(f"Error fetching IAM policy for bucket {bucket_name}: {e}")
                    continue
                is_public = any(
                    binding.get("role", "") in [
                        "roles/storage.objectViewer", "roles/storage.legacyBucketReader", "roles/storage.legacyObjectReader"
                    ] and "allUsers" in binding.get("members", []) for binding in policy.get("bindings", [])
                )
                if is_public:
                    bucket_data.append({
                        "Bucket Name": bucket_name,
                        "Location": bucket_location,
                        "Storage Class": storage_class,
                        "Access": "Public to internet",
                        "Comments from Ankercloud": "Restrict access to specific users or roles",
                        f"Comments from {client_name}": ""
                    })
            except Exception as e:
                print(f"Error processing bucket {bucket}: {e}")
                continue
    except Exception as e:
        print(f"Error listing buckets in project {project_id}: {e}")

    columns = ["S No", "Bucket Name", "Location", "Storage Class", "Access", "Comments from Ankercloud", f"Comments from {client_name}"]
    if not bucket_data:
        return pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(bucket_data)
        df.index += 1
        df.reset_index(inplace=True)
        df.rename(columns={"index": "S No"}, inplace=True)
        return df

# ============================
# 4. XLSX WRITING LOGIC
# ============================

TITLE_FILL = PatternFill(fill_type="solid", fgColor="FFCCEFFF")
HEADER_FONT = Font(name="Lato", bold=True, size=12)
TITLE_FONT = Font(name="Lato", bold=True, size=14)
BODY_FONT = Font(name="Lato", size=11)
THIN_SIDE = Side(border_style="thin", color="D0D7DE")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def sanitize_sheet_title(value):
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join(ch for ch in value if ch not in invalid_chars).strip()
    return (cleaned or "Sheet")[:31]


def _apply_border(ws, start_row, end_row, max_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def write_section_to_worksheet(ws, section_title, df, start_row):
    columns = list(df.columns) if df is not None else []
    num_columns = len(columns) if columns else 6
    title_row = start_row
    header_row = start_row + 1
    data_row = start_row + 2

    ws.cell(row=title_row, column=1, value=section_title)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=num_columns)
    title_cell = ws.cell(row=title_row, column=1)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = TITLE_FILL

    if columns:
        for col_idx, header in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        ws.cell(row=header_row, column=1, value="No columns")

    if df is None or df.empty:
        ws.cell(row=data_row, column=1, value="No action required")
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=num_columns)
        msg_cell = ws.cell(row=data_row, column=1)
        msg_cell.font = BODY_FONT
        msg_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        end_row = data_row
    else:
        for row_offset, row in enumerate(df.values.tolist(), start=0):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=data_row + row_offset, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        end_row = data_row + len(df) - 1

    _apply_border(ws, title_row, end_row, num_columns)
    return end_row + 2


def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            if value:
                max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 60))


# ============================
# 5. MAIN LOGIC (CONCURRENT)
# ============================


def audit_project(project_name, project_id, client_name, creds):
    crm, iam, compute, storage = get_gcp_services(project_id, creds)
    print(f"Auditing {project_name} ({project_id})...")

    report = {}
    report['IAM Users'] = fetch_iam_users(project_id, crm, client_name)
    report['IAM Groups'] = fetch_iam_groups(project_id, crm, client_name)
    report['Service Accounts'] = fetch_service_accounts(project_id, client_name, crm, iam)
    report['Firewall Rules'] = fetch_firewall_rules(project_id, compute, client_name)
    report['Public Buckets'] = fetch_public_buckets(project_id, storage, client_name)
    return report


def build_workbook(project_order, reports):
    workbook = Workbook()
    sheet_created = False
    used_titles = set()

    for display_name in project_order:
        sections = reports.get(display_name)
        if sections is None:
            continue
        base_title = sanitize_sheet_title(display_name)
        sheet_title = base_title
        counter = 2
        while sheet_title in used_titles:
            suffix = f"-{counter}"
            trimmed = base_title[: max(0, 31 - len(suffix))]
            sheet_title = f"{trimmed}{suffix}" if trimmed else f"Sheet{suffix}"
            counter += 1
        used_titles.add(sheet_title)

        if not sheet_created:
            ws = workbook.active
            ws.title = sheet_title
            sheet_created = True
        else:
            ws = workbook.create_sheet(title=sheet_title)

        start_row = 1
        for section_title, df in sections.items():
            start_row = write_section_to_worksheet(ws, section_title, df, start_row)

        autosize_columns(ws)

    if not sheet_created:
        raise ValueError("No audit data available to export.")
    return workbook


def generate_gcp_security_audit_xlsx(service_key, project_ids=None):
    creds = build_credentials(service_key)
    projects = resolve_project_selection(creds, project_ids)

    project_order = []
    project_map = []
    for project in projects:
        project_id = project.get("project_id")
        display_name = project.get("display_name") or project_id
        if display_name and display_name != project_id:
            display_label = f"{display_name} ({project_id})"
        else:
            display_label = project_id or "unknown-project"
        project_order.append(display_label)
        project_map.append((display_label, project_id))

    reports = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        future_to_project = {
            executor.submit(audit_project, display, pid, display, creds): display
            for display, pid in project_map
        }
        for future in concurrent.futures.as_completed(future_to_project):
            display_name = future_to_project[future]
            try:
                reports[display_name] = future.result()
            except Exception as exc:
                print(f"Failed for {display_name}: {exc}")

    workbook = build_workbook(project_order, reports)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime('%b-%Y')
    filename = f"GCP-Security-Audit-{timestamp}.xlsx"

    return {"filename": filename, "content": buffer.getvalue()}


def parse_cli_args():
    parser = argparse.ArgumentParser(description="Generate GCP security audit report as XLSX.")
    parser.add_argument("--service-key-file", dest="service_key_file", help="Path to GCP service account JSON.")
    parser.add_argument("--projects", dest="projects", help="Comma-separated GCP project IDs to audit.")
    parser.add_argument("--output", dest="output", help="Output XLSX filename.")
    return parser.parse_args()


def main():
    args = parse_cli_args()
    service_key = None
    if args.service_key_file:
        service_key = Path(args.service_key_file).read_text(encoding="utf-8")
    else:
        env_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
        if env_path and Path(env_path).exists():
            service_key = Path(env_path).read_text(encoding="utf-8")

    if not service_key:
        print("GCP service account JSON is required. Provide --service-key-file or set GOOGLE_APPLICATION_CREDENTIALS.")
        return

    project_ids = []
    if args.projects:
        project_ids = [item.strip() for item in args.projects.split(",") if item.strip()]
    result = generate_gcp_security_audit_xlsx(service_key, project_ids=project_ids)
    output_path = args.output or result["filename"]
    Path(output_path).write_bytes(result["content"])
    print(f"Report saved to {output_path}")


if __name__ == "__main__":
    main()
