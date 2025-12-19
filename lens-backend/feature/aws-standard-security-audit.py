import argparse
import os
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from zoneinfo import ZoneInfo

import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ================== CONFIGURATION ==================
BILLING_ACCOUNT_ID = "661907950294"
ROLE_NAME_DEFAULT = "securityaudit-automation"

PER_SERVICE_WORKERS = 5


def discover_accounts(base_session, role_name=ROLE_NAME_DEFAULT):
    try:
        org_client = base_session.client("organizations")
        org_meta = org_client.describe_organization().get("Organization", {})
        management_id = org_meta.get("ManagementAccountId") or org_meta.get("MasterAccountId")
        paginator = org_client.get_paginator("list_accounts")
        accounts = []
        for page in paginator.paginate():
            for acct in page.get("Accounts", []):
                if acct.get("Status") != "ACTIVE":
                    continue
                acct_id = acct.get("Id")
                name = acct.get("Name") or acct_id
                entry = {"id": acct_id, "name": name}
                if acct_id and acct_id != management_id:
                    entry["role"] = role_name
                accounts.append(entry)
        if accounts:
            return accounts
    except Exception as exc:
        print(f"Unable to list organization accounts: {exc}. Falling back to current account.")

    try:
        sts = base_session.client("sts")
        acct_id = sts.get_caller_identity().get("Account")
    except Exception:
        acct_id = None
    return [{"id": acct_id or "unknown", "name": "Current Account"}]


# ================== AWS SESSION HELPERS ==================
def assume_role(account_id, role_name, base_session, session_name="audit-session"):
    """
    Assume the specified IAM role for the given AWS account and return a session.
    """
    sts = base_session.client("sts")
    role_arn = f"arn:aws:iam::{account_id}:role/{role_name}"
    response = sts.assume_role(RoleArn=role_arn, RoleSessionName=session_name)
    creds = response['Credentials']
    session = boto3.Session(
        aws_access_key_id=creds['AccessKeyId'],
        aws_secret_access_key=creds['SecretAccessKey'],
        aws_session_token=creds['SessionToken']
    )
    return session

def get_session_for_account(acct, base_session):
    """
    Return a boto3 session for the account, assuming role if specified.
    """
    if 'role' in acct:
        return assume_role(acct['id'], acct['role'], base_session)
    else:
        return base_session

# ================== AWS AUDIT LOGIC (MAJOR TABLES) ===================
def get_all_aws_regions(session):
    ec2 = session.client('ec2', region_name='us-east-1')
    try:
        regions = ec2.describe_regions(AllRegions=True)
        return [region['RegionName'] for region in regions['Regions']
                if region['OptInStatus'] in ['opt-in-not-required', 'opted-in']]
    except Exception as e:
        print(f"Could not enumerate regions: {e}")
        return []

def is_bucket_public(bucket_name, s3_client):
    try:
        bucket_policy = s3_client.get_bucket_policy(Bucket=bucket_name)
        import json
        policy = json.loads(bucket_policy['Policy'])
        for statement in policy['Statement']:
            if statement.get('Effect') == 'Allow':
                principal = statement.get('Principal')
                if principal == "*" or (isinstance(principal, dict) and principal.get("AWS") == "*"):
                    actions = statement.get('Action', [])
                    if isinstance(actions, str):
                        actions = [actions]
                    if "s3:GetObject" in actions or "*" in actions:
                        return True
    except ClientError as e:
        if e.response['Error']['Code'] not in ['NoSuchBucketPolicy', 'AccessDenied']:
            print(f"Error checking policy for bucket {bucket_name}: {e}")
    try:
        bucket_acl = s3_client.get_bucket_acl(Bucket=bucket_name)
        for grant in bucket_acl['Grants']:
            grantee = grant.get('Grantee', {})
            uri = grantee.get('URI', '')
            if uri in [
                "http://acs.amazonaws.com/groups/global/AllUsers",
                "http://acs.amazonaws.com/groups/global/AuthenticatedUsers"
            ]:
                if grant.get('Permission') in ["READ", "FULL_CONTROL"]:
                    return True
    except Exception as e:
        print(f"Error checking ACL for bucket {bucket_name}: {e}")
    return False

def get_bucket_region(bucket_name, s3_client):
    try:
        location = s3_client.get_bucket_location(Bucket=bucket_name)
        return location.get('LocationConstraint') or 'us-east-1'
    except ClientError as e:
        if e.response['Error']['Code'] == 'AccessDenied':
            return 'Access Denied'
        print(f"Error getting region for bucket {bucket_name}: {e}")
        return 'Unknown'
    except Exception as e:
        print(f"Error getting region for bucket {bucket_name}: {e}")
        return 'Unknown'

def format_region(region_code):
    region_mapping = {
        'us-east-1': 'US East (N. Virginia)',
        'us-west-1': 'US West (N. California)',
        'us-west-2': 'US West (Oregon)',
        'eu-west-1': 'EU (Ireland)',
        'eu-central-1': 'EU (Frankfurt)',
        'ap-south-1': 'Asia Pacific (Mumbai)',
        'ap-southeast-1': 'Asia Pacific (Singapore)',
        'ap-northeast-1': 'Asia Pacific (Tokyo)',
    }
    return f"{region_mapping.get(region_code, 'Unknown')} ({region_code})"

def format_creation_date(date):
    """
    Format the creation date in a human-readable format (Asia/Kolkata timezone).
    """
    local_tz = ZoneInfo("Asia/Kolkata")
    local_time = date.astimezone(local_tz)
    return local_time.strftime('%B %d, %Y, %H:%M:%S')

def scan_single_bucket(args):
    """
    Scan a single S3 bucket for public access and return its status row.
    """
    bucket, index, session = args
    s3_client = session.client('s3')
    bucket_name = bucket['Name']
    try:
        creation_date = format_creation_date(bucket['CreationDate'])
        region_code = get_bucket_region(bucket_name, s3_client)
        region = format_region(region_code)
        try:
            objects = s3_client.list_objects_v2(Bucket=bucket_name, MaxKeys=1)
            is_empty = 'Contents' not in objects or len(objects['Contents']) == 0
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDenied':
                return None
            else:
                print(f"Error listing objects in bucket {bucket_name}: {e}")
                return None
        except Exception as e:
            print(f"Error listing objects in bucket {bucket_name}: {e}")
            return None

        is_public = is_bucket_public(bucket_name, s3_client)
        if is_empty and is_public:
            access = 'Public'
            comment_anker = "Review: Empty public bucket detected. Please confirm if deletion is appropriate."
        elif is_public:
            access = 'Public'
            comment_anker = "Public access policy detected. Restrict access to authorized users only."
        else:
            return None

        row = [
            index,
            bucket_name,
            region,
            creation_date,
            access,
            comment_anker
        ]
        row.append("")  # Customer comments
        for i in range(len(row)-1):
            if not row[i] or str(row[i]).strip() == "":
                row[i] = "-"
        return row
    except Exception as e:
        print(f"Error processing bucket {bucket_name}: {e}")
        return None

def get_s3_bucket_status_table_concurrent(session, max_workers=PER_SERVICE_WORKERS):
    """
    Scan all S3 buckets concurrently and return headers and data rows.
    """
    s3_client = session.client('s3')
    try:
        response = s3_client.list_buckets()
    except Exception as e:
        print(f"Error listing buckets: {e}")
        return ([
            'S.No',
            'Bucket Name',
            'AWS Region',
            'Creation Time',
            'Access (Public or private or both)',
            'Comment from Ankercloud',
            'Comments from Customer'
        ], [])
    bucket_list = list(enumerate(response['Buckets'], start=1))
    bucket_data = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(scan_single_bucket, (bucket, i, session)): (bucket, i) for i, bucket in bucket_list}
        for future in as_completed(futures):
            row = future.result()
            if row:
                bucket_data.append(row)
    for idx, row in enumerate(sorted(bucket_data, key=lambda x: x[1]), 1):
        row[0] = idx
    headers = [
        'S.No',
        'Bucket Name',
        'AWS Region',
        'Creation Time',
        'Access (Public or private or both)',
        'Comment from Ankercloud',
        'Comments from Customer'
    ]
    return headers, bucket_data

def scan_volumes_in_region(region, session):
    data = []
    try:
        ec2_client = session.client('ec2', region_name=region)
        response = ec2_client.describe_volumes()
        idx = 1
        for volume in response['Volumes']:
            if not volume['Encrypted']:
                volume_id = volume['VolumeId']
                volume_type = volume['VolumeType']
                size = f"{volume['Size']} GiB"
                iops = volume.get('Iops', '-')
                throughput = volume.get('Throughput', '-')
                snapshot_id = volume.get('SnapshotId', '-')
                creation_date = volume['CreateTime']
                formatted_creation_date = creation_date.strftime('%Y/%m/%d %H:%M GMT+5:30')
                availability_zone = volume['AvailabilityZone']
                volume_state = volume['State']
                name_tag = '-'
                if 'Tags' in volume:
                    for tag in volume['Tags']:
                        if tag['Key'] == 'Name':
                            name_tag = tag['Value']
                            break
                row = [
                    idx,
                    name_tag,
                    volume_id,
                    volume_type,
                    size,
                    iops,
                    throughput,
                    snapshot_id,
                    formatted_creation_date,
                    availability_zone,
                    volume_state,
                    region,
                    'Not encrypted',
                    'Enable encryption for this EBS volume to meet security best practices.'
                ]
                row.append("")  # Customer comments
                for i in range(len(row)-1):
                    if not row[i] or str(row[i]).strip() == "":
                        row[i] = "-"
                data.append(row)
                idx += 1
    except Exception as e:
        print(f"Skipping region {region}: {e}")
    return data

def get_unencrypted_volumes_table_all_regions_concurrent(session):
    regions = get_all_aws_regions(session)
    results = []
    idx = 1
    with ThreadPoolExecutor(max_workers=PER_SERVICE_WORKERS) as executor:
        futures = {executor.submit(scan_volumes_in_region, region, session): region for region in regions}
        for future in as_completed(futures):
            result = future.result()
            if result:
                for row in result:
                    row[0] = idx
                    idx += 1
                results.extend(result)
    headers = [
        'S.No',
        'Name',
        'Volume ID',
        'Type',
        'Size',
        'IOPS',
        'Throughput',
        'Snapshot',
        'Created',
        'Availability Zone',
        'Volume State',
        'Region',
        'Encryption',
        'Comment from Ankercloud',
        'Comments from Customer'
    ]
    return headers, results

def get_attached_instances(ec2, sg_id):
    instances = []
    try:
        response = ec2.describe_instances(
            Filters=[
                {'Name': 'instance.group-id', 'Values': [sg_id]}
            ]
        )
        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                instances.append(instance['InstanceId'])
    except Exception:
        pass
    return instances

def scan_secgroups_in_region(region, session):
    data = []
    try:
        ec2 = session.client('ec2', region_name=region)
        response = ec2.describe_security_groups()
        idx = 1
        for sg in response['SecurityGroups']:
            sg_id = sg.get('GroupId', '-')
            sg_name = sg.get('GroupName', '-')
            attached_instances = get_attached_instances(ec2, sg_id)
            for permission in sg['IpPermissions']:
                for ip_range in permission.get('IpRanges', []):
                    if ip_range.get('CidrIp') == '0.0.0.0/0':
                        from_port = permission.get('FromPort', -1)
                        to_port = permission.get('ToPort', -1)
                        if (from_port == 80 and to_port == 80) or (from_port == 443 and to_port == 443):
                            continue
                        if from_port == -1 or to_port == -1:
                            port_range = "All"
                        else:
                            port_range = f"{from_port}-{to_port}" if from_port != to_port else str(from_port)
                        row = [
                            idx,
                            sg_id,
                            sg_name,
                            ', '.join(attached_instances) if attached_instances else "-",
                            port_range,
                            ip_range['CidrIp'],
                            region,
                            'Restrict open ports to trusted IPs. Avoid exposing non-standard ports to the public internet.'
                        ]
                        row.append("")  # Customer comments
                        for i in range(len(row)-1):
                            if not row[i] or str(row[i]).strip() == "":
                                row[i] = "-"
                        data.append(row)
                        idx += 1
                        break
    except Exception as e:
        print(f"Skipping region {region}: {e}")
    return data

def get_security_groups_table_all_regions_concurrent(session):
    regions = get_all_aws_regions(session)
    results = []
    idx = 1
    with ThreadPoolExecutor(max_workers=PER_SERVICE_WORKERS) as executor:
        futures = {executor.submit(scan_secgroups_in_region, region, session): region for region in regions}
        for future in as_completed(futures):
            result = future.result()
            if result:
                for row in result:
                    row[0] = idx
                    idx += 1
                results.extend(result)
    headers = [
        'S.No', 'Security Group ID', 'Security Group Name', 'Attached Instances',
        'Ports', 'Listen', 'Region', 'Comment from Ankercloud', 'Comments from Customer'
    ]
    return headers, results

def is_mfa_enabled(iam_client, user_name):
    response = iam_client.list_mfa_devices(UserName=user_name)
    return len(response['MFADevices']) > 0

def get_user_access_keys(iam_client, user_name):
    response = iam_client.list_access_keys(UserName=user_name)
    has_access_key = False
    key_age = None
    last_activity = None
    if response['AccessKeyMetadata']:
        has_access_key = True
        for key_metadata in response['AccessKeyMetadata']:
            key_create_date = key_metadata['CreateDate']
            now = datetime.now()
            key_age = (now - key_create_date.replace(tzinfo=None)).days
    return {'has_access_key': has_access_key, 'key_age': key_age, 'last_activity': None}

def has_console_access(iam_client, user_name):
    try:
        iam_client.get_login_profile(UserName=user_name)
        return True
    except iam_client.exceptions.NoSuchEntityException:
        return False

def get_iam_users_table(session):
    try:
        iam_client = session.client('iam')
        paginator = iam_client.get_paginator('list_users')
        response_iterator = paginator.paginate()
        user_data = []
        serial_number = 1
        for page in response_iterator:
            for user in page['Users']:
                user_name = user['UserName']
                mfa_status = "Enabled" if is_mfa_enabled(iam_client, user_name) else "Disabled"
                access_keys = get_user_access_keys(iam_client, user_name)
                login_profile = has_console_access(iam_client, user_name)
                console_access = "Enabled" if login_profile else "Disabled"
                comment_anker = "Status: No action required."
                if console_access == "Enabled" and mfa_status == "Disabled" and access_keys['key_age'] is not None and access_keys['key_age'] > 90:
                    comment_anker = "Enable MFA and rotate access keys older than 90 days."
                elif console_access == "Enabled" and mfa_status == "Disabled":
                    comment_anker = "Enable MFA for users with console access."
                elif access_keys['key_age'] is not None and access_keys['key_age'] > 90:
                    comment_anker = "Rotate access keys older than 90 days."
                row = [
                    serial_number,
                    user_name if user_name else "-",
                    "-",  # Last Activity (not tracked in this example)
                    "-",  # Password Age
                    console_access,
                    "Yes" if access_keys['has_access_key'] else "Disabled",
                    f"{access_keys['key_age']} days" if access_keys['key_age'] is not None else "-",
                    mfa_status,
                    comment_anker
                ]
                row.append("")  # Comments from Customer
                for i in range(len(row)-1):
                    if not row[i] or str(row[i]).strip() == "":
                        row[i] = "-"
                user_data.append(row)
                serial_number += 1
        headers = [
            'S.No.',
            'User Name',
            'Last Activity',
            'Password Age',
            'Console Access',
            'Programmatic Access',
            'Access Key Age',
            'MFA Status',
            'Comment from Ankercloud',
            'Comments from Customer'
        ]
        return headers, user_data
    except Exception as e:
        print(f"Skipping IAM Users Table: {type(e).__name__}: {e}")
        return [
            'S.No.',
            'User Name',
            'Last Activity',
            'Password Age',
            'Console Access',
            'Programmatic Access',
            'Access Key Age',
            'MFA Status',
            'Comment from Ankercloud',
            'Comments from Customer'
        ], []


# ================== XLSX HELPERS ==================
TITLE_FILL = PatternFill(fill_type="solid", fgColor="FFCCEFFF")
HEADER_FONT = Font(name="Lato", bold=True, size=12)
TITLE_FONT = Font(name="Lato", bold=True, size=14)
BODY_FONT = Font(name="Lato", size=11)
THIN_SIDE = Side(border_style="thin", color="D0D7DE")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def sanitize_sheet_title(value):
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join(ch for ch in value if ch not in invalid_chars).strip()
    return (cleaned or "Sheet")[:31]


def _apply_border(ws, start_row, end_row, max_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def write_section_to_worksheet(ws, section_title, headers, rows, start_row):
    if rows:
        num_columns = len(headers)
        title_row = start_row
        header_row = start_row + 1
        data_start = start_row + 2
        data_end = data_start + len(rows) - 1

        ws.cell(row=title_row, column=1, value=section_title)
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=num_columns)
        title_cell = ws.cell(row=title_row, column=1)
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = TITLE_FILL

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_offset, row in enumerate(rows, start=0):
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=data_start + row_offset, column=col, value=value)
                cell.font = BODY_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        _apply_border(ws, title_row, data_end, num_columns)
        return data_end + 3

    num_columns = 4
    title_row = start_row
    message_row = start_row + 1
    ws.cell(row=title_row, column=1, value=section_title)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=num_columns)
    title_cell = ws.cell(row=title_row, column=1)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = TITLE_FILL

    ws.cell(row=message_row, column=1, value="No action required")
    ws.merge_cells(start_row=message_row, start_column=1, end_row=message_row, end_column=num_columns)
    msg_cell = ws.cell(row=message_row, column=1)
    msg_cell.font = BODY_FONT
    msg_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _apply_border(ws, title_row, message_row, num_columns)
    return message_row + 3


def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            if value:
                max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 60))


# ================== MAIN AUDIT LOOP ==================

def _build_billing_session(access_key, secret_key, session_token=None):
    if not access_key or not secret_key:
        raise ValueError("AWS access key and secret key are required.")
    return boto3.Session(
        aws_access_key_id=access_key,
        aws_secret_access_key=secret_key,
        aws_session_token=session_token,
    )


def audit_account(acct, base_session):
    session = get_session_for_account(acct, base_session)
    acct_name = acct["name"]
    print(f"Auditing {acct_name} ...")
    s3_headers, s3_rows = get_s3_bucket_status_table_concurrent(session)
    iam_headers, iam_rows = get_iam_users_table(session)
    sec_headers, sec_rows = get_security_groups_table_all_regions_concurrent(session)
    ebs_headers, ebs_rows = get_unencrypted_volumes_table_all_regions_concurrent(session)
    tables = [
        ("S3 Bucket Status", s3_headers, s3_rows),
        ("IAM Users Status", iam_headers, iam_rows),
        ("Security Groups Open to World (except 80/443)", sec_headers, sec_rows),
        ("Unencrypted EBS Volumes", ebs_headers, ebs_rows),
    ]
    return acct_name, tables


def audit_accounts_concurrent(accounts, base_session, max_workers=5):
    all_data = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(audit_account, acct, base_session): acct for acct in accounts}
        for future in as_completed(futures):
            acct = futures[future]
            try:
                acct_name, tables = future.result()
                all_data[acct_name] = tables
            except Exception as exc:
                name = acct.get("name") if isinstance(acct, dict) else "Account"
                print(f"Failed for {name}: {exc}")
    return all_data


def build_workbook(all_data):
    workbook = Workbook()
    sheet_created = False
    used_titles = set()

    for acct_name, tables in all_data.items():
        base_title = sanitize_sheet_title(acct_name)
        sheet_title = base_title
        counter = 2
        while sheet_title in used_titles:
            suffix = f"-{counter}"
            trimmed = base_title[: max(0, 31 - len(suffix))]
            sheet_title = f"{trimmed}{suffix}" if trimmed else f"Sheet{suffix}"
            counter += 1
        used_titles.add(sheet_title)

        if not sheet_created:
            ws = workbook.active
            ws.title = sheet_title
            sheet_created = True
        else:
            ws = workbook.create_sheet(title=sheet_title)

        start_row = 1
        for section_title, headers, rows in tables:
            start_row = write_section_to_worksheet(ws, section_title, headers, rows, start_row)

        autosize_columns(ws)

    if not sheet_created:
        raise ValueError("No audit data available to export.")
    return workbook


def generate_security_audit_xlsx(access_key, secret_key, session_token=None, role_name=ROLE_NAME_DEFAULT):
    base_session = _build_billing_session(access_key, secret_key, session_token)
    accounts = discover_accounts(base_session, role_name=role_name)
    all_data = audit_accounts_concurrent(accounts, base_session)
    workbook = build_workbook(all_data)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime('%b-%Y')
    filename = f"Security-Audit-{timestamp}.xlsx"

    return {"filename": filename, "content": buffer.getvalue()}


def parse_cli_args():
    parser = argparse.ArgumentParser(description="Generate AWS standard security audit report as XLSX.")
    parser.add_argument("--access-key", dest="access_key", help="AWS access key ID.")
    parser.add_argument("--secret-key", dest="secret_key", help="AWS secret access key.")
    parser.add_argument("--session-token", dest="session_token", help="Optional AWS session token.")
    parser.add_argument("--role-name", dest="role_name", help="Role name to assume in member accounts.")
    parser.add_argument("--output", dest="output", help="Output XLSX filename.")
    return parser.parse_args()


def main():
    args = parse_cli_args()
    access_key = args.access_key or os.environ.get("AWS_ACCESS_KEY_ID")
    secret_key = args.secret_key or os.environ.get("AWS_SECRET_ACCESS_KEY")
    session_token = args.session_token or os.environ.get("AWS_SESSION_TOKEN")
    role_name = args.role_name or ROLE_NAME_DEFAULT

    if not access_key or not secret_key:
        print("AWS access key and secret key are required.")
        return

    result = generate_security_audit_xlsx(
        access_key=access_key,
        secret_key=secret_key,
        session_token=session_token,
        role_name=role_name,
    )

    output_path = args.output or result["filename"]
    with open(output_path, "wb") as handle:
        handle.write(result["content"])
    print(f"Report saved to {output_path}")


if __name__ == "__main__":
    main()
