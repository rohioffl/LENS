import argparse
import json
import os
import re
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from getpass import getpass
from io import BytesIO

import boto3
from botocore.exceptions import ClientError

REGION_PATTERN = re.compile(r"^[a-z]{2}(?:-gov)?-[a-z0-9-]+-\d+$")
PLACEHOLDER_VALUES = {"", "n/a", "na", "none", "-", "null"}

RESOURCE_MAP = {
    1: 'cost',
    2: 'rds',
    3: 'elasticache',
    4: 'backup',
    5: 'secretsmanager',
    6: 'appsync',
    7: 'dynamodb',
    8: 'cloudwatch',
    9: 'ecs',
    10: 'kms',
    11: 'mq',
    12: 'codecommit',
    13: 'codepipeline',
    14: 'ecr',
    15: 'codebuild',
    16: 'codeartifact',
    17: 'cloudformation',
    18: 'waf',
    19: 'eks',
    20: 'codedeploy',
    21: 'vpc',
    22: 'iam_identity',
    23: 'ec2',
    24: 'redshift',
    25: 'sqs',
    26: 'stepfunctions',
    27: 'route53',
    28: 'sns',
    29: 'lambda',
    30: 'glue',
    31: 'efs',
    32: 'amplify',
    33: 'cloudfront',
    34: 's3',
    35: 'iam_user',
    36: 'iam_group',
    37: 'iam_policies',
    38: 'iam_role',
}

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - runtime guard for missing dependency
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None

    def get_column_letter(_col):  # pragma: no cover - fallback to keep name defined
        raise RuntimeError("openpyxl is required for spreadsheet export")


class ListWriter:
    """Lightweight writer that mimics csv.writer for in-memory row capture."""

    def __init__(self):
        self.rows = []

    def writerow(self, row):
        # Ensure row is always a list to keep openpyxl happy
        if not isinstance(row, (list, tuple)):
            row = [row]

        safe_row = []
        for value in row:
            if value is None:
                safe_row.append("")
            elif isinstance(value, datetime):
                if value.tzinfo is not None:
                    value = value.astimezone(timezone.utc).replace(tzinfo=None)
                safe_row.append(value)
            elif isinstance(value, date):
                safe_row.append(value)
            elif isinstance(value, Decimal):
                safe_row.append(float(value))
            elif isinstance(value, (list, tuple, set)):
                try:
                    safe_row.append(json.dumps(list(value), default=str))
                except TypeError:
                    safe_row.append(str(value))
            elif isinstance(value, dict):
                try:
                    safe_row.append(json.dumps(value, default=str))
                except TypeError:
                    safe_row.append(str(value))
            elif isinstance(value, bytes):
                safe_row.append(value.decode("utf-8", errors="replace"))
            else:
                safe_row.append(value)

        self.rows.append(safe_row)


def sanitize_sheet_name(name):
    """Trim sheet names to Excel constraints and strip unsupported characters."""
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join(ch for ch in name if ch not in invalid_chars).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:31]


def parse_relative_date_range(value):
    """Return (start_iso, end_iso) if value matches 'last N days' pattern."""
    if not value:
        return None

    normalized = re.sub(r"[-_]+", " ", value.lower().strip())
    tokens = normalized.split()
    if len(tokens) < 2 or tokens[0] != "last":
        return None

    quantity = None
    unit = None
    for token in tokens[1:]:
        if quantity is None and token.isdigit():
            quantity = int(token)
        elif unit is None and token.startswith("day"):
            unit = token

    if quantity is None:
        return None

    if unit is None:
        unit = "days"
    elif not unit.startswith("day"):
        return None

    today = datetime.utcnow().date()
    start_date = today - timedelta(days=quantity)
    end_date = today
    return start_date.isoformat(), end_date.isoformat()


def normalize_manual_date(value):
    if not value:
        return ""

    cleaned = value.strip()
    if not cleaned:
        return ""

    known_formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
        "%Y.%m.%d",
        "%m-%d-%Y",
        "%m/%d/%Y",
    ]

    for fmt in known_formats:
        try:
            parsed = datetime.strptime(cleaned, fmt).date()
            return parsed.isoformat()
        except ValueError:
            continue

    # If no format matched, return the original string for downstream handling
    return cleaned


def resolve_cost_period(start_input, end_input):
    """Normalize manual or relative cost period inputs into ISO date strings."""
    relative = parse_relative_date_range(start_input)
    if relative:
        return relative

    relative = parse_relative_date_range(end_input)
    if relative:
        return relative

    start = start_input.strip()
    end = end_input.strip() if end_input else ""

    if not start and end:
        relative = parse_relative_date_range(end)
        if relative:
            return relative

    start = normalize_manual_date(start)
    end = normalize_manual_date(end)

    if not end:
        end = datetime.utcnow().date().isoformat()

    if not start:
        start = end

    return start, end


def stringify_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def table_block_has_data(rows, start_idx, end_idx):
    if start_idx > end_idx:
        return False

    for idx in range(start_idx, end_idx + 1):
        raw_values = [stringify_cell(v).strip() for v in rows[idx]]
        non_empty = [val for val in raw_values if val]
        if not non_empty:
            continue

        lower_vals = [val.lower() for val in non_empty]
        meaningful = []
        for original, lower in zip(non_empty, lower_vals):
            if lower in PLACEHOLDER_VALUES:
                continue
            if REGION_PATTERN.match(lower) or lower == "global":
                continue
            if "no resource found" in lower:
                continue
            meaningful.append(original)

        if meaningful:
            return True
    return False


def extract_table_blocks(rows):
    blocks = []
    start_idx = None
    header_len = 0
    max_cols = 0

    for idx, row in enumerate(rows):
        values = [stringify_cell(v).strip() for v in row]
        has_content = any(values)
        if len(row) > 1 and has_content:
            if start_idx is None:
                start_idx = idx
                header_len = len(row)
                max_cols = len(row)
            else:
                max_cols = max(max_cols, len(row))
        else:
            if start_idx is not None:
                end_idx = idx - 1
                if end_idx > start_idx and table_block_has_data(rows, start_idx + 1, end_idx):
                    blocks.append(
                        {
                            "start": start_idx,
                            "end": end_idx,
                            "header_len": header_len,
                            "max_cols": max_cols,
                        }
                    )
                start_idx = None
                header_len = 0
                max_cols = 0

    if start_idx is not None:
        end_idx = len(rows) - 1
        if end_idx > start_idx and table_block_has_data(rows, start_idx + 1, end_idx):
            blocks.append(
                {
                    "start": start_idx,
                    "end": end_idx,
                    "header_len": header_len,
                    "max_cols": max_cols,
                }
            )

    return blocks


def style_titles(sheet, rows):
    title_font = Font(name="Lato", bold=True, size=12)
    alignment = Alignment(horizontal="left", vertical="center")

    for idx, row in enumerate(rows, start=1):
        if len(row) == 1:
            value = stringify_cell(row[0]).strip()
            if value:
                cell = sheet.cell(row=idx, column=1)
                cell.font = title_font
                cell.alignment = alignment


def apply_table_formatting(sheet, rows, table_blocks):
    header_font = Font(name="Lato", bold=True, size=12)
    body_font = Font(name="Lato", bold=False, size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    band_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_side = Side(border_style="thin", color="D0D7DE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for block in table_blocks:
        start_row = block["start"] + 1
        end_row = block["end"] + 1
        header_len = block["header_len"]
        max_cols = block["max_cols"]

        # Header styling
        for col in range(1, header_len + 1):
            cell = sheet.cell(row=start_row, column=col)
            cell.font = header_font
            cell.alignment = alignment
            cell.fill = header_fill
            cell.border = border

        # Data rows
        for row_idx in range(start_row + 1, end_row + 1):
            fill = band_fill if (row_idx - start_row) % 2 == 0 else None
            for col in range(1, max_cols + 1):
                cell = sheet.cell(row=row_idx, column=col)
                cell.font = body_font
                cell.alignment = alignment
                cell.border = border
                if fill:
                    cell.fill = fill


def autofit_columns(sheet):
    for col_idx in range(1, sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in sheet[column_letter]:
            value = stringify_cell(cell.value)
            if not value:
                continue
            max_length = max(max_length, len(value))
        if max_length == 0:
            width = 12
        else:
            width = min(max_length + 2, 60)
        sheet.column_dimensions[column_letter].width = width


def parse_cli_args():
    parser = argparse.ArgumentParser(
        description="Generate AWS service inventories as XLSX workbooks.",
        epilog="Credential flags are optional; any missing value will be requested interactively unless a profile is supplied.",
    )
    parser.add_argument(
        "--access-key",
        dest="access_key",
        help="AWS access key ID to use for authentication. Requires --secret-key; prompted if omitted.",
    )
    parser.add_argument(
        "--secret-key",
        dest="secret_key",
        help="AWS secret access key to use for authentication. Requires --access-key; prompted if omitted.",
    )
    parser.add_argument(
        "--session-token",
        dest="session_token",
        help="Optional AWS session token for temporary credentials.",
    )
    parser.add_argument(
        "--profile",
        dest="profile_name",
        help="Optional AWS named profile (disables credential prompts). Ignored when explicit credentials are provided.",
    )
    parser.add_argument(
        "--regions",
        dest="regions",
        help="Comma-separated AWS regions to inspect (skips the interactive region prompt).",
    )
    parser.add_argument(
        "--resources",
        dest="resources",
        help="Comma-separated resource identifiers (numbers or names) or 'all' to skip the interactive menu.",
    )
    parser.add_argument(
        "--from-date",
        dest="from_date",
        help="Cost analysis start date (YYYY-MM-DD, DD-MM-YYYY, or relative text like 'last 30 days').",
    )
    parser.add_argument(
        "--to-date",
        dest="to_date",
        help="Cost analysis end date (YYYY-MM-DD or DD-MM-YYYY). Leave blank to default to today unless --from-date includes a relative range.",
    )
    return parser.parse_args()


def create_boto3_session(access_key=None, secret_key=None, session_token=None, profile_name=None):
    session_kwargs = {}
    explicit_creds = any([access_key, secret_key, session_token])

    if explicit_creds:
        if not access_key or not secret_key:
            raise ValueError("Both --access-key and --secret-key must be provided when specifying credentials.")
        session_kwargs["aws_access_key_id"] = access_key
        session_kwargs["aws_secret_access_key"] = secret_key
        if session_token:
            session_kwargs["aws_session_token"] = session_token
    elif profile_name:
        session_kwargs["profile_name"] = profile_name

    return boto3.session.Session(**session_kwargs)


def prompt_if_missing(value, prompt_text, secret=False):
    if value:
        return value
    responder = getpass if secret else input
    try:
        response = responder(prompt_text)
    except EOFError:
        return value
    response = response.strip()
    return response or None


def parse_regions_input(raw_value):
    if not raw_value:
        return []
    return [region.strip() for region in raw_value.split(',') if region.strip()]


def parse_resource_selection(raw_value, resource_map):
    if not raw_value:
        return None

    value = raw_value.strip().lower()
    if value == "all":
        return list(resource_map.values())

    selected_resources = []
    tokens = [token.strip() for token in value.split(',') if token.strip()]

    for token in tokens:
        if token.isdigit():
            number = int(token)
            if number in resource_map:
                selected_resources.append(resource_map[number])
        else:
            normalized = token.replace(" ", "_")
            if normalized in resource_map.values():
                selected_resources.append(normalized)

    return selected_resources or None


def resolve_credentials(args):
    access_key = args.access_key
    secret_key = args.secret_key
    session_token = args.session_token
    profile_name = args.profile_name

    if profile_name:
        return access_key, secret_key, session_token, profile_name

    access_key = prompt_if_missing(
        access_key,
        "Enter AWS access key ID (press Enter to use the default credential chain): ",
    )

    if access_key or secret_key or session_token:
        if not access_key:
            access_key = prompt_if_missing(
                access_key,
                "AWS access key ID is required when providing secret credentials. Enter AWS access key ID: ",
            )
        if not access_key:
            raise ValueError("AWS access key ID is required when supplying secret credentials.")

        secret_key = prompt_if_missing(
            secret_key,
            "Enter AWS secret access key: ",
            secret=True,
        )
        if not secret_key:
            raise ValueError("AWS secret access key is required when supplying secret credentials.")
    else:
        access_key = None
        secret_key = None

    return access_key, secret_key, session_token, profile_name


def resolve_cost_period_inputs(from_date_arg, to_date_arg):
    if not from_date_arg and not to_date_arg:
        return None

    if from_date_arg:
        relative = parse_relative_date_range(from_date_arg)
        if relative:
            return relative

    from_date = from_date_arg or ""
    to_date = to_date_arg or ""
    return resolve_cost_period(from_date, to_date)

# Function to fetch tags as a formatted string
def get_tags(tags):
    return ", ".join([f"{tag['Key']}={tag['Value']}" for tag in tags]) if isinstance(tags, list) else 'N/A'

# Function to fetch all tags as a string
def get_all_tags(tags):
    return ", ".join([f"{tag['Key']}:{tag['Value']}" for tag in tags]) if isinstance(tags, list) else "N/A"

# Function to fetch instance name from tags
def get_instance_name(tags):
    for tag in tags:
        if tag['Key'] == 'Name':
            return tag['Value']
    return 'N/A'

# Function to get creation date of resources, updated with check for 'CreationTime'
def get_creation_date(obj):
    if 'CreationTime' in obj:
        return obj['CreationTime'].strftime('%Y-%m-%d %H:%M:%S')
    else:
        return 'N/A'

# Function to fetch tags using Resource Group Tagging API
def fetch_tags(resourcegroupstaggingapi_client, resource_arn):
    try:
        response = resourcegroupstaggingapi_client.get_resources(
            ResourceARNList=[resource_arn]
        )
        if response.get('ResourceTagMappingList'):
            tags = response['ResourceTagMappingList'][0].get('Tags', [])
            return get_tags(tags)
        else:
            return "N/A"
    except ClientError as e:
        print(f"Error fetching tags for {resource_arn}: {e}")
        return "N/A"


# -------------------------- IAM Group Inventory --------------------------
def get_iam_group_inventory(iam_client, resourcegroupstaggingapi_client, writer):
    writer.writerow(['IAM Group Inventory'])
    writer.writerow(['GroupName', 'GroupId', 'CreateDate', 'Tags', 'Users'])  # Added 'CreateDate' column

    try:
        groups = iam_client.list_groups().get('Groups', [])
        for group in groups:
            group_name = group['GroupName']
            group_id = group['GroupId']
            create_date = group['CreateDate'].strftime('%Y-%m-%d %H:%M:%S') if 'CreateDate' in group else 'N/A'

            # Fetch group tags using Resource Group Tagging API
            group_arn = f"arn:aws:iam::{iam_client.meta.region_name}:group/{group_name}"
            tags = get_tags(group_arn)

            # Get users in the group
            users_in_group = []
            try:
                users = iam_client.get_group(GroupName=group_name).get('Users', [])
                users_in_group = [user['UserName'] for user in users]
            except ClientError:
                users_in_group = []

            writer.writerow([group_name, group_id, create_date, tags, ", ".join(users_in_group)])
    except ClientError as e:
        print(f"Error fetching IAM groups: {e}")

    writer.writerow([])  # Blank row to separate sections

# -------------------------- IAM Policy Inventory --------------------------
def get_iam_policy_inventory(iam_client, resourcegroupstaggingapi_client, writer):
    writer.writerow(['IAM Policy Inventory'])
    writer.writerow(['PolicyName', 'PolicyArn', 'CreateDate', 'Tags', 'AttachedEntities'])

    try:
        policies = iam_client.list_policies(Scope='All').get('Policies', [])
        for policy in policies:
            policy_name = policy.get('PolicyName', 'N/A')
            policy_arn = policy.get('Arn', 'N/A')
            create_date = policy.get('CreateDate')
            create_date_str = create_date.strftime('%Y-%m-%d %H:%M:%S') if create_date else 'N/A'

            # Fetch tags
            tags = fetch_tags(resourcegroupstaggingapi_client, policy_arn)

            # Fetch attached entities
            attached_entities = []
            try:
                entities = iam_client.list_entities_for_policy(PolicyArn=policy_arn)
                attached_entities = [
                    *[group['GroupName'] for group in entities.get('PolicyGroups', [])],
                    *[user['UserName'] for user in entities.get('PolicyUsers', [])],
                    *[role['RoleName'] for role in entities.get('PolicyRoles', [])],
                ]
            except ClientError as e:
                print(f"Error fetching entities for {policy_name}: {e}")
                attached_entities = []

            writer.writerow([policy_name, policy_arn, create_date_str, tags, ", ".join(attached_entities)])
    except ClientError as e:
        print(f"Error fetching IAM policies: {e}")
        writer.writerow(['Error fetching IAM policies', '', '', '', ''])

    writer.writerow([])  # Blank row to separate sections


# -------------------------- IAM Role Inventory --------------------------
def get_iam_role_inventory(iam_client, resourcegroupstaggingapi_client, writer):
    writer.writerow(['IAM Role Inventory'])
    writer.writerow(['RoleName', 'RoleArn', 'AssumeRolePolicyDocument', 'CreateDate', 'Tags'])  # Added 'CreateDate' column

    try:
        roles = iam_client.list_roles().get('Roles', [])
        for role in roles:
            role_name = role['RoleName']
            role_arn = role['Arn']
            assume_role_policy_document = json.dumps(role['AssumeRolePolicyDocument'], default=str)
            create_date = role['CreateDate'].strftime('%Y-%m-%d %H:%M:%S') if 'CreateDate' in role else 'N/A'

            # Fetch role tags using Resource Group Tagging API
            tags = get_tags(role_arn)

            writer.writerow([role_name, role_arn, assume_role_policy_document, create_date, tags])
    except ClientError as e:
        print(f"Error fetching IAM roles: {e}")

    writer.writerow([])  # Blank row to separate sections


# -------------------------- IAM User Inventory --------------------------
def get_iam_user_inventory(iam_client, resourcegroupstaggingapi_client, writer):
    writer.writerow(['IAM User Inventory'])
    writer.writerow(['UserName', 'UserArn', 'CreateDate', 'Tags'])  # Added 'CreateDate' column

    try:
        users = iam_client.list_users().get('Users', [])
        for user in users:
            user_name = user['UserName']
            user_arn = user['Arn']
            create_date = user['CreateDate'].strftime('%Y-%m-%d %H:%M:%S') if 'CreateDate' in user else 'N/A'

            # Fetch user tags using Resource Group Tagging API
            user_arn = f"arn:aws:iam::{iam_client.meta.region_name}:user/{user_name}"
            tags = get_tags(user_arn)

            writer.writerow([user_name, user_arn, create_date, tags])
    except ClientError as e:
        print(f"Error fetching IAM users: {e}")

    writer.writerow([])  # Blank row to separate sections

# Function to generate CSV for specified resource type
def generate_inventory_csv(regions, resource_type, writer, session, cost_period=None, interactive=True):
    # Initialize inventory list
    inventory = []

    # -------------------------- Cost Report --------------------------
    if resource_type.lower() == 'cost':
        writer.writerow(['Cost-Related Resources'])
        writer.writerow(['Service', 'Unblended Cost (USD)', 'Blended Cost (USD)', 'Total Cost (USD)'])

        ce_client = session.client('ce')  # AWS Cost Explorer client

        if cost_period:
            start_date, end_date = cost_period
        elif interactive:
            start_date = input("Enter the start date for cost analysis (YYYY-MM-DD): ")
            end_date = input("Enter the end date for cost analysis (YYYY-MM-DD): ")
        else:
            raise ValueError("cost_period is required when requesting cost data without interactive prompts.")

        try:
            response = ce_client.get_cost_and_usage(
                TimePeriod={'Start': start_date, 'End': end_date},
                Granularity='MONTHLY',
                Metrics=['BlendedCost', 'UnblendedCost'],
                GroupBy=[{'Type': 'DIMENSION', 'Key': 'SERVICE'}],
            )

            # Debug: show available metrics for the first service
            if response['ResultsByTime'] and response['ResultsByTime'][0]['Groups']:
                print("Available metrics:", response['ResultsByTime'][0]['Groups'][0]['Metrics'])

            for result in response['ResultsByTime']:
                groups = result.get('Groups', [])
                for group in groups:
                    service = group['Keys'][0] if group['Keys'] else 'N/A'
                    unblended_amount = float(group['Metrics'].get('UnblendedCost', {}).get('Amount', 0))
                    blended_amount = float(group['Metrics'].get('BlendedCost', {}).get('Amount', 0))
                    total_cost = unblended_amount  # you could also use blended

                    # ✅ Only write rows if service incurred cost > 0
                    if total_cost > 0:
                        writer.writerow([
                            service,
                            f"${unblended_amount:.2f}",
                            f"${blended_amount:.2f}",
                            f"${total_cost:.2f}"
                        ])

        except ClientError as e:
            print(f"Error fetching cost data: {str(e)}")
            writer.writerow(['Error fetching cost data', '', '', ''])

        writer.writerow([])  # blank line after cost section

    # -------------------------- RDS Inventory --------------------------
    elif resource_type.lower() == 'rds':
        writer.writerow(['RDS Inventory'])
        writer.writerow(['Region', 'DBInstanceIdentifier', 'Engine', 'Status', 'RetentionPeriod',
                     'ParameterGroup', 'Storage AutoScaling', 'ReplicationLag', 'Endpoint', 'Storage (GB)'])

        for region in regions:
            client = session.client('rds', region_name=region)
            try:
                instances = client.describe_db_instances().get('DBInstances', [])
                for instance in instances:
                    storage = instance.get('AllocatedStorage', 'N/A')
                    replication_source = instance.get('ReadReplicaSourceDBInstanceIdentifier', None)
                
                    # Check if the instance is a read replica and has replication lag
                    if replication_source:
                        replication_lag = instance.get('ReplicationSourceIdentifier', 'N/A')
                    else:
                        replication_lag = 'N/A'

                    writer.writerow([region,
                                    instance['DBInstanceIdentifier'],
                                    instance['Engine'],
                                    instance['DBInstanceStatus'],
                                    instance.get('BackupRetentionPeriod', 'N/A'),
                                    instance.get('DBParameterGroups', [{'DBParameterGroupName': 'N/A'}])[0]['DBParameterGroupName'],
                                    'Enabled' if instance.get('AutoMinorVersionUpgrade') else 'Disabled',
                                    replication_lag,
                                    instance.get('Endpoint', {}).get('Address', 'N/A'),
                                    storage])
            except ClientError as e:
                print(f"Error fetching RDS instances in {region}: {e}")

        writer.writerow([])
        

    #---------------------------Elastic Cahe Inventory---------------------
    elif resource_type.lower() == 'elasticache':

        writer.writerow(['ElastiCache Clusters Inventory'])
        fieldnames_clusters = ['Region', 'ClusterId', 'Engine', 'EngineVersion', 'CacheNodeType', 'NumCacheNodes', 'Tag']
        writer.writerow(fieldnames_clusters)

        inventory_clusters = []

        for region in regions:
            client = session.client('elasticache', region_name=region)
            try:
                clusters = client.describe_cache_clusters()['CacheClusters']
                if not clusters:
                    inventory_clusters.append({
                        'Region': region, 
                        'ClusterId': 'No resource found', 
                        'Engine': 'N/A', 
                        'EngineVersion': 'N/A', 
                        'CacheNodeType': 'N/A', 
                        'NumCacheNodes': 'N/A', 
                        'Tag': 'N/A'
                    })
                else:
                    for cluster in clusters:
                        inventory_clusters.append({
                            'Region': region,
                            'ClusterId': cluster['CacheClusterId'],
                            'Engine': cluster['Engine'],
                            'EngineVersion': cluster['EngineVersion'],
                            'CacheNodeType': cluster['CacheNodeType'],
                            'NumCacheNodes': cluster['NumCacheNodes'],
                            'Tag': 'N/A'
                        })
            except ClientError as e:
                print(f"Error fetching ElastiCache clusters in {region}: {e}")

        # Write collected data
        for row in inventory_clusters:
            writer.writerow([row['Region'], row['ClusterId'], row['Engine'], row['EngineVersion'], row['CacheNodeType'], row['NumCacheNodes'], row['Tag']])

        writer.writerow([])  # Blank row to separate sections

        writer.writerow(['ElastiCache Backups Inventory'])
        fieldnames_backups = ['Region', 'Name', 'CacheClusterId', 'CacheNodeType', 'Source', 'Status', 'CreateTime', 'Tag']
        writer.writerow(fieldnames_backups)

        inventory_backups = []

        for region in regions:
            client = session.client('elasticache', region_name=region)
            try:
                snapshots = client.describe_snapshots()['Snapshots']
                if not snapshots:
                    inventory_backups.append({
                        'Region': region, 
                        'Name': 'No resource found', 
                        'CacheClusterId': 'N/A', 
                        'CacheNodeType': 'N/A', 
                        'Source': 'N/A', 
                        'Status': 'N/A', 
                        'CreateTime': 'N/A', 
                        'Tag': 'N/A'
                    })
                else:
                    for snapshot in snapshots:
                        inventory_backups.append({
                            'Region': region,
                            'Name': snapshot['SnapshotName'],
                            'CacheClusterId': snapshot.get('CacheClusterId', 'N/A'),
                            'CacheNodeType': snapshot.get('CacheNodeType', 'N/A'),
                            'Source': snapshot.get('SnapshotSource', 'N/A'),
                            'Status': snapshot['SnapshotStatus'],
                            'CreateTime': snapshot['SnapshotCreateTime'].strftime('%Y-%m-%d %H:%M:%S') if snapshot.get('SnapshotCreateTime') else 'N/A',
                            'Tag': 'N/A'
                        })
            except ClientError as e:
                print(f"Error fetching ElastiCache backups in {region}: {e}")

        # Write collected data
        for row in inventory_backups:
            writer.writerow([row['Region'], row['Name'], row['CacheClusterId'], row['CacheNodeType'], row['Source'], row['Status'], row['CreateTime'], row['Tag']])
            
             
        writer.writerow([])  # Blank row to separate sections
    
# -------------------------- Backup Inventory --------------------------
    elif resource_type.lower() == 'backup':

        writer.writerow(['Backup Inventory'])
        # Write Backup Plan header
        writer.writerow(['Region', 'Type', 'BackupPlanId', 'BackupPlanName', 'CreationDate', 'LastExecutionDate', 'VersionId'])

        for region in regions:  # Changed from REGIONS to regions
            backup_client = session.client('backup', region_name=region)
            region_resources_found = False

            # Fetch Backup Plans
            try:
                plans = backup_client.list_backup_plans().get('BackupPlansList', [])
                if plans:
                    region_resources_found = True
                    for plan in plans:
                        plan_id = plan['BackupPlanId']
                        plan_name = plan['BackupPlanName']
                        creation_date = plan['CreationDate'].strftime('%Y-%m-%d %H:%M:%S')
                        last_execution_date = plan.get('LastExecutionDate', 'N/A')
                        if last_execution_date != 'N/A':
                            last_execution_date = last_execution_date.strftime('%Y-%m-%d %H:%M:%S')
                        version_id = plan.get('VersionId', 'N/A')

                        writer.writerow([region, 'Backup Plan', plan_id, plan_name, creation_date, last_execution_date, version_id])

            except ClientError as e:
                print(f"Error fetching Backup Plans in {region}: {str(e)}")

            if not region_resources_found:
                writer.writerow([region, 'No resource found', '', '', '', '', ''])

        # Write a blank line before the Backup Vault section
        writer.writerow([])

        # Write Backup Vault header
        writer.writerow(['Region', 'Type', 'BackupVaultName', 'CreationDate', 'EncryptionKeyArn', 'NumberOfRecoveryPoints'])

        region_resources_found = False

        # Fetch Backup Vaults
        for region in regions:  # Changed from REGIONS to regions
            try:
                vaults = backup_client.list_backup_vaults().get('BackupVaultList', [])
                if vaults:
                    region_resources_found = True
                    for vault in vaults:
                        vault_name = vault['BackupVaultName']
                        creation_date = vault['CreationDate'].strftime('%Y-%m-%d %H:%M:%S')
                        encryption_key = vault.get('EncryptionKeyArn', 'N/A')
                        recovery_points = vault.get('NumberOfRecoveryPoints', 0)

                        writer.writerow([region, 'Backup Vault', vault_name, creation_date, encryption_key, recovery_points])

            except ClientError as e:
                print(f"Error fetching Backup Vaults in {region}: {str(e)}")

            if not region_resources_found:
                writer.writerow([region, 'No resource found', '', '', '', ''])

        writer.writerow([])  # Blank row to separate sections

# -------------------------- Secret Manager Inventory --------------------------
    elif resource_type.lower() == 'secretsmanager':

        writer.writerow(['Secret Manager Inventory'])
        writer.writerow(['Region', 'SecretId**', 'Name', 'Description', 'LastChangedDate', 'Tags'])
        
        for region in regions:
            secret_client = session.client('secretsmanager', region_name=region)
            region_resources_found = False
            try:
                secrets = secret_client.list_secrets()['SecretList']
                if secrets:
                    region_resources_found = True
                    for secret in secrets:
                        secret_id = secret['ARN']
                        name = secret['Name']
                        description = secret.get('Description', 'N/A')
                        last_changed = secret['LastChangedDate'].strftime('%Y-%m-%d %H:%M:%S') if 'LastChangedDate' in secret else 'N/A'
                        tags = secret.get('Tags', 'N/A')
                        writer.writerow([region, secret_id, name, description, last_changed, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])

            except ClientError as e:
                print(f"Error fetching secrets in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- AppSync Inventory --------------------------
    elif resource_type.lower() == 'appsync':

        writer.writerow(['AppSync Inventory'])
        writer.writerow(['Region', 'APIId', 'Name', 'AuthenticationType', 'CreatedAt', 'Tags', 'APIKeyExpirationDates'])
        
        for region in regions:
            appsync_client = session.client('appsync', region_name=region)
            region_resources_found = False
            try:
                apis = appsync_client.list_graphql_apis().get('graphqlApis', [])
                if apis:
                    region_resources_found = True
                    for api in apis:
                        api_id = api.get('apiId', 'N/A')
                        name = api.get('name', 'N/A')
                        auth_type = api.get('authenticationType', 'N/A')

                        # Handle createdAt safely
                        created_at = api.get('createdAt')
                        if isinstance(created_at, int):  # If it's a timestamp
                            created_at = datetime.fromtimestamp(created_at, tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
                        elif hasattr(created_at, 'strftime'):  # If it's a datetime object
                            created_at = created_at.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            created_at = 'N/A'

                        tags = api.get('tags', 'N/A')

                        # Fetch API keys and their expiration dates
                        try:
                            api_keys = appsync_client.list_api_keys(apiId=api_id).get('apiKeys', [])
                            expiration_dates = ', '.join(
                                datetime.fromtimestamp(key['expires'], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
                                if isinstance(key['expires'], int)
                                else key['expires'].strftime('%Y-%m-%d %H:%M:%S')
                                for key in api_keys if 'expires' in key
                            )
                            expiration_dates = expiration_dates if expiration_dates else 'No API Keys'
                        except ClientError as e:
                            expiration_dates = f"Error: {e}"

                        writer.writerow([region, api_id, name, auth_type, created_at, tags, expiration_dates])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A', ''])
            except ClientError as e:
                print(f"Error fetching AppSync APIs in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- DynamoDB Inventory --------------------------
    elif resource_type.lower() == 'dynamodb':

        writer.writerow(['DynamoDB Inventory'])
        writer.writerow(['Region', 'TableName', 'TableStatus', 'CreationDate', 'ProvisionedThroughput', 'ItemCount', 'Tags'])
        
        for region in regions:
            dynamodb_client = session.client('dynamodb', region_name=region)
            region_resources_found = False
            try:
                tables = dynamodb_client.list_tables()['TableNames']
                if tables:
                    region_resources_found = True
                    for table in tables:
                        table_info = dynamodb_client.describe_table(TableName=table)['Table']
                        table_status = table_info['TableStatus']
                        creation_date = table_info['CreationDateTime'].strftime('%Y-%m-%d %H:%M:%S')
                        throughput = table_info.get('ProvisionedThroughput', {})
                        provisioned_throughput = f"Read: {throughput.get('ReadCapacityUnits', 'N/A')}, Write: {throughput.get('WriteCapacityUnits', 'N/A')}"
                        item_count = table_info.get('ItemCount', 'N/A')
                        tags = dynamodb_client.list_tags_of_resource(ResourceArn=table_info['TableArn'])['Tags'] if 'TableArn' in table_info else 'N/A'
                        writer.writerow([region, table, table_status, creation_date, provisioned_throughput, item_count, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching DynamoDB tables in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- CloudWatch Inventory --------------------------
    elif resource_type.lower() == 'cloudwatch':

        writer.writerow(['Cloudwatch Inventory'])

        # -------------------------- CloudWatch Alarms Section --------------------------
        writer.writerow(['Region', 'AlarmName', 'StateValue', 'MetricName', 'Namespace', 'ActionsEnabled'])

        for region in regions:
            cloudwatch_client = session.client('cloudwatch', region_name=region)
            region_resources_found = False
            try:
                alarms = cloudwatch_client.describe_alarms()['MetricAlarms']
                if alarms:
                    region_resources_found = True
                    for alarm in alarms:
                        alarm_name = alarm['AlarmName']
                        state_value = alarm['StateValue']
                        metric_name = alarm['MetricName']
                        namespace = alarm['Namespace']
                        actions_enabled = alarm.get('ActionsEnabled', 'False')
                        writer.writerow([region, alarm_name, state_value, metric_name, namespace, actions_enabled])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching CloudWatch alarms in {region}: {e}")
        
        # -------------------------- CloudWatch Log Groups Section --------------------------
        writer.writerow([])  # Add a blank row between Alarms and Log Groups sections
        writer.writerow(['Region', 'LogGroupName', 'CreationDate', 'RetentionPeriod', 'Tags'])

        for region in regions:
            logs_client = session.client('logs', region_name=region)
            region_resources_found = False

            try:
                log_groups = logs_client.describe_log_groups().get('logGroups', [])
                if log_groups:
                    region_resources_found = True
                    for log_group in log_groups:
                        name = log_group['logGroupName']
                        creation_ts = log_group['creationTime'] / 1000
                        creation_dt = datetime.fromtimestamp(creation_ts, tz=timezone.utc)
                        creation_date = creation_dt.strftime('%Y-%m-%d %H:%M:%S')
                        retention_period = log_group.get('retentionInDays', 'Never Expire')
                        tags = logs_client.list_tags_log_group(logGroupName=name).get('tags', 'N/A')
                        writer.writerow([region, name, creation_date, retention_period, tags])

                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', 'N/A'])

            except ClientError as e:
                print(f"Error fetching CloudWatch Log Groups in {region}: {e}")

        writer.writerow([])  # Blank row to separate sections

# -------------------------- ECS Inventory --------------------------
    elif resource_type.lower() == 'ecs':

        writer.writerow(['ECS Inventory'])
        writer.writerow(['Region', 'ClusterName', 'Status', 'RunningTasksCount', 'Tags'])
        
        for region in regions:
            ecs_client = session.client('ecs', region_name=region)
            region_resources_found = False
            try:
                clusters = ecs_client.list_clusters()['clusterArns']
                if clusters:
                    region_resources_found = True
                    for cluster in clusters:
                        cluster_name = cluster.split('/')[-1]
                        cluster_info = ecs_client.describe_clusters(clusters=[cluster_name])['clusters'][0]
                        status = cluster_info['status']
                        running_tasks = cluster_info['runningTasksCount']
                        tags = ecs_client.list_tags_for_resource(resourceArn=cluster_info['clusterArn'])['tags'] if 'clusterArn' in cluster_info else 'N/A'
                        writer.writerow([region, cluster_name, status, running_tasks, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching ECS clusters in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- KMS Inventory --------------------------
    elif resource_type.lower() == 'kms':

        writer.writerow(['kms Inventory'])
        writer.writerow(['Region', 'KeyId', 'Description', 'Enabled', 'CreationDate', 'Tags'])
        
        for region in regions:
            kms_client = session.client('kms', region_name=region)
            region_resources_found = False
            try:
                # List the keys and iterate over them
                keys = kms_client.list_keys()['Keys']
                if keys:
                    region_resources_found = True
                    for key_id in keys:
                        # Ensure the key_id is extracted as a string
                        key_id_str = key_id if isinstance(key_id, str) else key_id.get('KeyId', 'N/A')
                        key_info = kms_client.describe_key(KeyId=key_id_str)['KeyMetadata']
                        description = key_info['Description']
                        enabled = key_info['Enabled']
                        creation_date = key_info['CreationDate'].strftime('%Y-%m-%d %H:%M:%S')
                        
                        # Fetch tags (if present)
                        try:
                            tags = kms_client.list_resource_tags(KeyId=key_id_str)['Tags']
                        except ClientError:
                            tags = 'N/A'

                        writer.writerow([region, key_id_str, description, enabled, creation_date, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching KMS keys in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections


# -------------------------- MQ Inventory --------------------------
    elif resource_type.lower() == 'mq':

        writer.writerow(['MQ Inventory'])
        writer.writerow(['Region', 'BrokerName', 'BrokerArn', 'Status', 'EngineType', 'Tags'])
        
        for region in regions:
            mq_client = session.client('mq', region_name=region)
            region_resources_found = False
            try:
                brokers = mq_client.list_brokers()['BrokerSummaries']
                if brokers:
                    region_resources_found = True
                    for broker in brokers:
                        broker_name = broker['BrokerName']
                        broker_arn = broker['BrokerArn']
                        status = broker['BrokerState']
                        engine_type = broker['EngineType']
                        tags = mq_client.list_tags(ResourceArn=broker_arn).get('Tags', 'N/A')
                        writer.writerow([region, broker_name, broker_arn, status, engine_type, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching MQ brokers in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- Codecommit Inventory --------------------------
    elif resource_type.lower() == 'codecommit':

        writer.writerow(['Codecommit Inventory'])
        writer.writerow(['Region', 'RepositoryName', 'RepositoryArn', 'CreatedAt', 'LastModified', 'Tags'])
        
        for region in regions:
            codecommit_client = session.client('codecommit', region_name=region)
            region_resources_found = False
            try:
                repositories = codecommit_client.list_repositories()['repositories']
                if repositories:
                    region_resources_found = True
                    for repo in repositories:
                        repo_name = repo['repositoryName']
                        
                        # Fetch repository details
                        repo_details = codecommit_client.get_repository(repositoryName=repo_name)['repositoryMetadata']
                        repo_arn = repo_details['Arn']
                        creation_date = repo_details['creationDate'].strftime('%Y-%m-%d %H:%M:%S')
                        last_modified = (
                            repo_details.get('lastModifiedDate', 'N/A').strftime('%Y-%m-%d %H:%M:%S') 
                            if 'lastModifiedDate' in repo_details else 'N/A'
                        )
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, repo_name, repo_arn, creation_date, last_modified, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching CodeCommit repositories in {region}: {e}")
                writer.writerow([region, 'Error', '', '', '', 'N/A'])
        
        writer.writerow([])  # Blank row to separate sections


# -------------------------- CodePipeline Inventory --------------------------
    elif resource_type.lower() == 'codepipeline':

        writer.writerow(['Codepipeline Inventory'])
        writer.writerow(['Region', 'PipelineName', 'PipelineArn', 'Created', 'LastModified', 'Tags'])
        
        for region in regions:
            codepipeline_client = session.client('codepipeline', region_name=region)
            region_resources_found = False
            try:
                pipelines = codepipeline_client.list_pipelines()['pipelines']
                if pipelines:
                    region_resources_found = True
                    for pipeline in pipelines:
                        pipeline_name = pipeline['name']
                        pipeline_arn = pipeline.get('arn', 'N/A')  # Safely accessing 'arn'
                        created = pipeline.get('created', 'N/A')
                        last_modified = pipeline.get('lastModified', 'N/A')
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, pipeline_name, pipeline_arn, created, last_modified, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching CodePipeline pipelines in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- ECR Inventory --------------------------
    elif resource_type.lower() == 'ecr':

        writer.writerow(['ECR Inventory'])
        writer.writerow(['Region', 'RepositoryName', 'RepositoryUri', 'Created', 'Tags'])
        
        for region in regions:
            ecr_client = session.client('ecr', region_name=region)
            region_resources_found = False
            try:
                repositories = ecr_client.describe_repositories()['repositories']
                if repositories:
                    region_resources_found = True
                    for repo in repositories:
                        repo_name = repo['repositoryName']
                        repo_uri = repo['repositoryUri']
                        created = repo.get('createdAt', 'N/A')
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, repo_name, repo_uri, created, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching ECR repositories in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- CodeBuild Inventory --------------------------
    elif resource_type.lower() == 'codebuild':

        writer.writerow(['CodeBuild Inventory'])
        writer.writerow(['Region', 'ProjectName', 'Arn', 'Created', 'LastModified', 'Environment', 'Tags'])
        
        for region in regions:
            codebuild_client = session.client('codebuild', region_name=region)
            region_resources_found = False
            try:
                projects = codebuild_client.list_projects()['projects']
                if projects:
                    region_resources_found = True
                    for project in projects:
                        project_info = codebuild_client.batch_get_projects(names=[project])['projects'][0]
                        project_name = project_info['name']
                        project_arn = project_info['arn']
                        created = project_info.get('created', 'N/A')
                        last_modified = project_info.get('lastModified', 'N/A')
                        environment = project_info['environment']['image']
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, project_name, project_arn, created, last_modified, environment, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching CodeBuild projects in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections


# -------------------------- CodeArtifact Inventory --------------------------
    elif resource_type.lower() == 'codeartifact':

        writer.writerow(['Code Artifact Inventory'])
        writer.writerow(['Region', 'RepositoryName', 'DomainName', 'RepositoryArn', 'Created', 'Tags'])
        
        for region in regions:
            codeartifact_client = session.client('codeartifact', region_name=region)
            region_resources_found = False
            try:
                repositories = codeartifact_client.list_repositories()['repositories']
                if repositories:
                    region_resources_found = True
                    for repo in repositories:
                        repo_name = repo['repositoryName']
                        domain_name = repo['domainName']
                        repo_arn = repo['repositoryArn']
                        created = repo.get('createdTime', 'N/A')  # Ensure correct field for created time
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, repo_name, domain_name, repo_arn, created, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching CodeArtifact repositories in {region}: {e}")
            except Exception as e:
                print(f"Unexpected error in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections


# -------------------------- CloudFormation Inventory --------------------------
    elif resource_type.lower() == 'cloudformation':

        writer.writerow(['CloudFormation Inventory'])
        writer.writerow(['Region', 'StackName', 'StackId', 'CreationTime', 'StackStatus', 'Tags'])
        
        for region in regions:
            cloudformation_client = session.client('cloudformation', region_name=region)
            region_resources_found = False
            try:
                stacks = cloudformation_client.describe_stacks()['Stacks']
                if stacks:
                    region_resources_found = True
                    for stack in stacks:
                        stack_name = stack['StackName']
                        stack_id = stack['StackId']
                        creation_time = stack.get('CreationTime', 'N/A').strftime('%Y-%m-%d %H:%M:%S') if 'CreationTime' in stack else 'N/A'
                        stack_status = stack['StackStatus']
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, stack_name, stack_id, creation_time, stack_status, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching CloudFormation stacks in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections


# -------------------------- WAF Inventory --------------------------
    elif resource_type.lower() == 'waf':

        writer.writerow(['WAF Inventory'])
        
        # Write WebACL section header
        writer.writerow(['WebACL Inventory'])
        writer.writerow(['Region', 'WebAclName', 'WebAclId', 'Created', 'Tags'])  # Header for WebACLs

        for region in regions:
            waf_client = session.client('wafv2', region_name=region)
            webacl_found = False

            try:
                # Fetch WebACLs
                web_acls = waf_client.list_web_acls(Scope='REGIONAL')['WebACLs']
                if web_acls:
                    webacl_found = True
                    for acl in web_acls:
                        acl_name = acl['Name']
                        acl_id = acl['Id']
                        created = acl['CreationTime'].strftime('%Y-%m-%d %H:%M:%S')
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, acl_name, acl_id, created, tags])
                
                if not webacl_found:
                    writer.writerow([region, 'No WebACL resource found', '', '', 'N/A'])

            except ClientError as e:
                print(f"Error fetching WebACLs in {region}: {e}")
                writer.writerow([region, 'Error', '', '', 'N/A'])

        # Write IP Set section header
        writer.writerow([])  # Empty row for separation
        writer.writerow(['IP Set Inventory'])
        writer.writerow(['Region', 'IPSetName', 'IPSetId', 'IPAddresses', 'Tags'])  # Header for IP Sets

        for region in regions:
            waf_client = session.client('wafv2', region_name=region)
            ipset_found = False

            try:
                # Fetch IP Sets
                ip_sets = waf_client.list_ip_sets(Scope='REGIONAL')['IPSets']
                if ip_sets:
                    ipset_found = True
                    for ip_set in ip_sets:
                        ip_set_name = ip_set['Name']
                        ip_set_id = ip_set['Id']
                        ip_addresses = ', '.join(ip_set.get('Addresses', []))  # Collect IP addresses
                        writer.writerow([region, ip_set_name, ip_set_id, ip_addresses, 'N/A'])

                if not ipset_found:
                    writer.writerow([region, 'No IP Set resource found', '', '', 'N/A'])

            except ClientError as e:
                print(f"Error fetching IP Sets in {region}: {e}")
                writer.writerow([region, 'Error', '', '', 'N/A'])


        writer.writerow([])  # Blank row to separate sections


# -------------------------- EKS Inventory --------------------------
    elif resource_type.lower() == 'eks':

        writer.writerow(['EKS Inventory'])
        writer.writerow(['Region', 'ClusterName', 'ClusterArn', 'Created', 'Status', 'Tags'])
        
        for region in regions:
            eks_client = session.client('eks', region_name=region)
            region_resources_found = False
            try:
                clusters = eks_client.list_clusters()['clusters']
                if clusters:
                    region_resources_found = True
                    for cluster in clusters:
                        cluster_info = eks_client.describe_cluster(name=cluster)['cluster']
                        cluster_name = cluster_info['name']
                        cluster_arn = cluster_info['arn']
                        created = cluster_info.get('createdAt', 'N/A')
                        status = cluster_info['status']
                        tags = 'N/A'  # Tags can be fetched using list_tags_for_resource
                        writer.writerow([region, cluster_name, cluster_arn, created, status, tags])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', 'N/A'])
            except ClientError as e:
                print(f"Error fetching EKS clusters in {region}: {e}")
        
        writer.writerow([])  # Blank row to separate sections

    # -------------------------- CodeDeploy Inventory --------------------------
    elif resource_type.lower() == 'codedeploy':

        writer.writerow(['CodeDeploy Inventory'])
        writer.writerow(['Region', 'Deployment ID', 'Status', 'Application', 'Deployment Group', 
                         'Revision Location', 'Initiating Event', 'Start Time', 'End Time'])
        
        for region in regions:
            codedeploy_client = session.client('codedeploy', region_name=region)
            region_resources_found = False
            try:
                deployments = codedeploy_client.list_deployments()['deployments']
                if deployments:
                    region_resources_found = True
                    for deployment_id in deployments:
                        deployment_info = codedeploy_client.get_deployment(deploymentId=deployment_id)['deployment']
                        deployment_status = deployment_info.get('status', 'N/A')
                        application_name = deployment_info.get('applicationName', 'N/A')
                        deployment_group = deployment_info.get('deploymentGroupName', 'N/A')
                        revision_location = deployment_info.get('revision', {}).get('revisionLocation', 'N/A')
                        initiating_event = deployment_info.get('initiatedBy', 'N/A')
                        start_time = deployment_info.get('startTime', 'N/A')
                        end_time = deployment_info.get('endTime', 'N/A')

                        # Convert times to string if they're available
                        start_time = start_time.strftime('%Y-%m-%d %H:%M:%S') if isinstance(start_time, datetime) else start_time
                        end_time = end_time.strftime('%Y-%m-%d %H:%M:%S') if isinstance(end_time, datetime) else end_time

                        writer.writerow([region, deployment_id, deployment_status, application_name, 
                                         deployment_group, revision_location, initiating_event, start_time, end_time])
                
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', '', '', '', ''])
            except ClientError as e:
                print(f"Error fetching CodeDeploy deployments in {region}: {e}")
        
        writer.writerow([])
    
# -------------------------- VPC Inventory --------------------------
    elif resource_type.lower() == 'vpc':

        writer.writerow(['VPC Inventory'])

        
        for region in regions:
            ec2 = session.client('ec2', region_name=region)
            
            # Add region header
            writer.writerow([f'Region: {region}', '', '', '', '', '', ''])
            
            # VPCs
            vpcs = ec2.describe_vpcs()['Vpcs']
            writer.writerow(['', 'VPCs', '', '', '', '', ''])
            writer.writerow(['ID', 'Name', 'State', 'Creation Date', 'Flow Log Status', 'Tags'])
            if vpcs:
                for vpc in vpcs:
                    vpc_id = vpc['VpcId']
                    vpc_name = next((tag['Value'] for tag in vpc.get('Tags', []) if tag['Key'] == 'Name'), '')
                    vpc_creation_date = get_creation_date(vpc)
                    
                    flow_logs = ec2.describe_flow_logs(Filters=[{'Name': 'resource-id', 'Values': [vpc_id]}])['FlowLogs']
                    flow_log_status = ", ".join([f"{log['LogDestinationType']}:{log['FlowLogId']}" for log in flow_logs]) if flow_logs else 'No Flow Logs'
                    vpc_tags = ", ".join([f"{tag['Key']}={tag['Value']}" for tag in vpc.get('Tags', [])])
                    
                    writer.writerow([vpc_id, vpc_name, vpc['State'], vpc_creation_date, flow_log_status, vpc_tags])
            else:
                writer.writerow(['No resource found', '', '', '', '', '', ''])
            
            # Subnets
            subnets = ec2.describe_subnets()['Subnets']
            writer.writerow(['', 'Subnets', '', '', '', '', ''])
            writer.writerow(['Name', 'State', 'VPC', 'Available IPv4 Addresses', 'ID', 'Creation Date', 'Tags'])
            if subnets:
                for subnet in subnets:
                    subnet_name = next((tag['Value'] for tag in subnet.get('Tags', []) if tag['Key'] == 'Name'), '')
                    subnet_creation_date = get_creation_date(subnet)
                    subnet_tags = ", ".join([f"{tag['Key']}={tag['Value']}" for tag in subnet.get('Tags', [])])
                    writer.writerow([subnet_name, subnet['State'], subnet['VpcId'], subnet['AvailableIpAddressCount'], subnet['SubnetId'], subnet_creation_date, subnet_tags])
            else:
                writer.writerow(['No resource found', '', '', '', '', '', ''])
            
            # Route Tables
            route_tables = ec2.describe_route_tables()['RouteTables']
            writer.writerow(['', 'Route Tables', '', '', '', '', ''])
            writer.writerow(['Name', 'State', 'VPC', 'Route Table ID', 'Route Destinations', 'Tags'])
            if route_tables:
                for route_table in route_tables:
                    route_table_name = next((tag['Value'] for tag in route_table.get('Tags', []) if tag['Key'] == 'Name'), '')
                    route_table_state = 'Main' if any(assoc.get('Main', False) for assoc in route_table['Associations']) else 'Custom'
                    route_table_vpc = route_table['VpcId'] if 'VpcId' in route_table else 'N/A'
                    route_destinations = ", ".join([route.get('DestinationCidrBlock', 'N/A') for route in route_table['Routes']])
                    route_table_tags = ", ".join([f"{tag['Key']}={tag['Value']}" for tag in route_table.get('Tags', [])])
                    writer.writerow([route_table_name, route_table_state, route_table_vpc, route_table['RouteTableId'], route_destinations, route_table_tags])
            else:
                writer.writerow(['No resource found', '', '', '', '', '', ''])
            
            # Internet Gateways
            internet_gateways = ec2.describe_internet_gateways()['InternetGateways']
            writer.writerow(['', 'Internet Gateways', '', '', '', '', ''])
            writer.writerow(['Name', 'State', 'VPC', 'ID', 'Tags'])
            if internet_gateways:
                for igw in internet_gateways:
                    igw_name = next((tag['Value'] for tag in igw.get('Tags', []) if tag['Key'] == 'Name'), '')
                    igw_state = igw['Attachments'][0]['State'] if igw['Attachments'] else 'N/A'
                    igw_vpc = igw['Attachments'][0]['VpcId'] if igw['Attachments'] else 'N/A'
                    igw_tags = ", ".join([f"{tag['Key']}={tag['Value']}" for tag in igw.get('Tags', [])])
                    writer.writerow([igw_name, igw_state, igw_vpc, igw['InternetGatewayId'], igw_tags])
            else:
                writer.writerow(['No resource found', '', '', '', '', ''])
            
            # NAT Gateways
            nat_gateways = ec2.describe_nat_gateways()['NatGateways']
            writer.writerow(['', 'NAT Gateways', '', '', '', '', ''])
            writer.writerow(['Name', 'State', 'VPC', 'Subnet', 'Creation Date', 'ID', 'Tags'])
            if nat_gateways:
                for nat_gw in nat_gateways:
                    nat_gw_name = next((tag['Value'] for tag in nat_gw.get('Tags', []) if tag['Key'] == 'Name'), '')
                    nat_gw_creation_date = get_creation_date(nat_gw)
                    nat_gw_tags = ", ".join([f"{tag['Key']}={tag['Value']}" for tag in nat_gw.get('Tags', [])])
                    writer.writerow([nat_gw_name, nat_gw['State'], nat_gw['VpcId'], nat_gw['SubnetId'], nat_gw_creation_date, nat_gw['NatGatewayId'], nat_gw_tags])
            else:
                writer.writerow(['No resource found', '', '', '', '', '', ''])
        
        writer.writerow([])  # Blank row to separate sections

# -------------------------- IAM Identity Center Inventory --------------------------
    elif resource_type.lower() == 'iam_identity':

        writer.writerow(['IAM Identity Center Inventory'])

        # Define field names for each section
        fieldnames_users = [
            'Username', 'User ID', 'Display name', 'Email', 'Status',
            'MFA devices', 'Group names', 'Created time', 'Last updated', 'Tags'
        ]
        fieldnames_groups = [
            'Group name', 'Group ID', 'Users', 'AWS Account',
            'Created time', 'Last updated', 'Tags'
        ]
        fieldnames_permission_sets = [
            'Permission set', 'ARN', 'Created date', 'Accounts',
            'Policies', 'Tags'
        ]

        # Initialize clients
        sso_admin_client = session.client('sso-admin')
        identitystore_client = session.client('identitystore')

        try:
            # Fetch the Identity Center instance
            instances = sso_admin_client.list_instances().get('Instances', [])
            if not instances:
                print("No Identity Center instances found.")
                writer.writerow(['No resource found'])
                return

            instance_arn = instances[0]['InstanceArn']
            identity_store_id = instances[0]['IdentityStoreId']

            # -------------------------- Fetch Users --------------------------
            writer.writerow([])
            writer.writerow(['Users'])
            writer.writerow(fieldnames_users)

            users = identitystore_client.list_users(IdentityStoreId=identity_store_id).get('Users', [])
            if not users:
                writer.writerow(['No resource found'] + ['N/A'] * (len(fieldnames_users) - 1))
            else:
                for user in users:
                    username = user.get('UserName', 'N/A')
                    user_id = user.get('UserId', 'N/A')
                    display_name = user.get('DisplayName', 'N/A')
                    email = user.get('Emails', [{}])[0].get('Value', 'N/A')
                    status = user.get('Status', 'N/A')
                    created_time = user.get('CreatedDate', 'N/A')
                    last_updated = user.get('LastUpdatedDate', 'N/A')
                    tags = get_tags(user.get('Tags', []))

                    # Placeholder for MFA details
                    mfa_devices_list = "MFA details not available in IdentityStore API"

                    # Fetch Group Names
                    group_memberships = identitystore_client.list_group_memberships_for_member(
                        IdentityStoreId=identity_store_id,
                        MemberId={'UserId': user_id}
                    ).get('GroupMemberships', [])
                    group_names = ', '.join([
                        identitystore_client.describe_group(
                            IdentityStoreId=identity_store_id,
                            GroupId=membership['GroupId']
                        ).get('Group', {}).get('DisplayName', 'N/A')
                        for membership in group_memberships
                    ]) or 'N/A'

                    writer.writerow([
                        username, user_id, display_name, email, status,
                        mfa_devices_list, group_names, created_time, last_updated, tags
                    ])

            # -------------------------- Fetch Groups --------------------------
            writer.writerow([])
            writer.writerow(['Groups'])
            writer.writerow(fieldnames_groups)

            groups = identitystore_client.list_groups(IdentityStoreId=identity_store_id).get('Groups', [])
            if not groups:
                writer.writerow(['No resource found'] + ['N/A'] * (len(fieldnames_groups) - 1))
            else:
                for group in groups:
                    group_name = group.get('DisplayName', 'N/A')
                    group_id = group.get('GroupId', 'N/A')
                    created_time = group.get('CreatedDate', 'N/A')
                    last_updated = group.get('LastUpdatedDate', 'N/A')
                    tags = get_tags(group.get('Tags', []))

                    # Fetch Users
                    group_memberships = identitystore_client.list_group_memberships(
                        IdentityStoreId=identity_store_id,
                        GroupId=group_id
                    ).get('GroupMemberships', [])
                    users_in_group = ', '.join([
                        identitystore_client.describe_user(
                            IdentityStoreId=identity_store_id,
                            UserId=membership['MemberId']['UserId']
                        ).get('UserName', 'N/A')
                        for membership in group_memberships
                    ]) or 'N/A'

                    writer.writerow([
                        group_name, group_id, users_in_group, 'N/A',  # AWS Account is optional, replace 'N/A' if needed
                        created_time, last_updated, tags
                    ])


        except ClientError as e:
            print(f"Error fetching IAM Identity Center data: {e}")
            
        writer.writerow([])  # Blank row to separate sections    


# -------------------------- EC2 Inventory --------------------------

    elif resource_type.lower() == 'ec2':
  
        for region in regions:
            ec2_client = session.client('ec2', region_name=region)
            elb_client = session.client('elb', region_name=region)
            elbv2_client = session.client('elbv2', region_name=region)
            autoscaling_client = session.client('autoscaling', region_name=region)
            
            writer.writerow([f'Region: {region}'])
            
            # Write headers for EC2 Instances
            writer.writerow(['EC2 Instances'])
            writer.writerow(['Instance ID', 'Instance Name', 'Instance Type', 'State', 'Public IP', 'Private IP', 'Launch Time', 'Security Groups', 'Tags'])
            
            # Fetch and write EC2 instances data
            instances = ec2_client.describe_instances()
            if not instances['Reservations']:
                writer.writerow(["No resources found"])
            for reservation in instances['Reservations']:
                for instance in reservation['Instances']:
                    security_groups = ", ".join([sg['GroupName'] for sg in instance['SecurityGroups']])
                    instance_name = get_instance_name(instance.get('Tags', []))
                    tags = get_all_tags(instance.get('Tags', []))
                    launch_time = get_creation_date(instance)  # Use updated get_creation_date function
                    writer.writerow([
                        instance['InstanceId'],
                        instance_name,
                        instance['InstanceType'],
                        instance['State']['Name'],
                        instance.get('PublicIpAddress', 'N/A'),
                        instance.get('PrivateIpAddress', 'N/A'),
                        launch_time,
                        security_groups,
                        tags
                    ])
            
            # Write headers for AMIs
            writer.writerow([])
            writer.writerow(['AMIs'])
            writer.writerow(['AMI ID', 'Name', 'State', 'Creation Date', 'Tags'])
            
            # Fetch and write AMIs data
            images = ec2_client.describe_images(Owners=['self'])
            if not images['Images']:
                writer.writerow(["No resources found"])
            for image in images['Images']:
                tags = get_all_tags(image.get('Tags', []))
                creation_date = get_creation_date(image)  # Use updated get_creation_date function
                writer.writerow([
                    image['ImageId'],
                    image.get('Name', 'N/A'),
                    image['State'],
                    creation_date,
                    tags
                ])
            
            # Write headers for Volumes
            writer.writerow([])
            writer.writerow(['Volumes'])
            writer.writerow(['Volume ID', 'Size (GiB)', 'State', 'Volume Type', 'Create Time', 'Attached Instances', 'Encrypted', 'Tags'])
            
            # Fetch and write Volumes data
            volumes = ec2_client.describe_volumes()['Volumes']
            if not volumes:
                writer.writerow(['No resource found'])
            for volume in volumes:
                attachments = ", ".join([attachment['InstanceId'] for attachment in volume['Attachments']])
                writer.writerow([
                    volume['VolumeId'],
                    volume['Size'],
                    volume['State'],
                    volume['VolumeType'],
                    get_creation_date(volume),  # Use updated get_creation_date function
                    attachments if attachments else 'N/A',
                    volume['Encrypted'],
                    get_tags(volume.get('Tags', []))
                ])
            
            # Write headers for Snapshots
            writer.writerow([])
            writer.writerow(['Snapshots'])
            writer.writerow(['Snapshot ID', 'Volume ID', 'State', 'Start Time', 'Tags'])
            
            # Fetch and write Snapshots data
            snapshots = ec2_client.describe_snapshots(OwnerIds=['self'])['Snapshots']
            if not snapshots:
                writer.writerow(['No resource found'])
            for snapshot in snapshots:
                writer.writerow([
                    snapshot['SnapshotId'],
                    snapshot['VolumeId'],
                    snapshot['State'],
                    snapshot['StartTime'],
                    get_tags(snapshot.get('Tags', []))
                ])
            
            # Write headers for Elastic IPs
            writer.writerow([])
            writer.writerow(['Elastic IPs'])
            writer.writerow(['Public IP', 'Allocation ID', 'Instance ID', 'Tags'])
            
            # Fetch and write Elastic IPs data
            eips = ec2_client.describe_addresses()['Addresses']
            if not eips:
                writer.writerow(['No resource found'])
            for address in eips:
                writer.writerow([
                    address['PublicIp'],
                    address.get('AllocationId', 'N/A'),
                    address.get('InstanceId', 'N/A'),
                    get_tags(address.get('Tags', []))
                ])
            
            # Write headers for Load Balancers
            writer.writerow([])
            writer.writerow(['Load Balancers'])
            writer.writerow(['Load Balancer Name', 'DNS Name', 'Type', 'Created Time', 'Tags'])
            
            # Fetch and write Load Balancers data
            load_balancers = elbv2_client.describe_load_balancers()['LoadBalancers']
            if not load_balancers:
                writer.writerow(['No resource found'])
            for lb in load_balancers:
                writer.writerow([
                    lb['LoadBalancerName'],
                    lb['DNSName'],
                    lb['Type'],
                    get_creation_date(lb),  # Use updated get_creation_date function
                    get_tags(lb.get('Tags', []))
                ])

            # Write headers for Target Groups
            writer.writerow([])
            writer.writerow(['Target Groups'])
            writer.writerow(['Target Group Name', 'Protocol', 'Port', 'Tags'])
            
            # Fetch and write Target Groups data
            target_groups = elbv2_client.describe_target_groups()['TargetGroups']
            if not target_groups:
                writer.writerow(['No resource found'])
            for tg in target_groups:
                writer.writerow([
                    tg['TargetGroupName'],
                    tg.get('Protocol', 'N/A'),
                    tg.get('Port', 'N/A'),
                    get_tags(tg.get('Tags', []))
                ])
            
            # Write headers for Auto Scaling Groups
            writer.writerow([])
            writer.writerow(['Auto Scaling Groups'])
            writer.writerow(['Auto Scaling Group Name', 'Min Size', 'Max Size', 'Desired Capacity', 'Created Time', 'Tags'])
            
            # Fetch and write Auto Scaling Groups data
            auto_scaling_groups = autoscaling_client.describe_auto_scaling_groups()['AutoScalingGroups']
            if not auto_scaling_groups:
                writer.writerow(['No resource found'])
            for asg in auto_scaling_groups:
                writer.writerow([
                    asg['AutoScalingGroupName'],
                    asg['MinSize'],
                    asg['MaxSize'],
                    asg['DesiredCapacity'],
                    get_creation_date(asg),  # Use updated get_creation_date function
                    get_tags(asg.get('Tags', []))
                ])
            
            # Write headers for Security Groups
            writer.writerow([])
            writer.writerow(['Security Groups'])
            writer.writerow(['Group ID', 'Group Name', 'Description', 'VPC ID', 'Tags'])
            
            # Fetch and write Security Groups data
            security_groups = ec2_client.describe_security_groups()['SecurityGroups']
            if not security_groups:
                writer.writerow(['No resource found'])
            for sg in security_groups:
                writer.writerow([
                    sg['GroupId'],
                    sg['GroupName'],
                    sg['Description'],
                    sg.get('VpcId', 'N/A'),
                    get_tags(sg.get('Tags', []))
                ])
            
            writer.writerow([])  # Add an empty row to separate regions

    # -------------------------- Red Shift--------------------------
    elif resource_type == 'redshift':
        writer.writerow(['Redshift Inventory'])
        writer.writerow(['Region', 'ClusterIdentifier', 'NodeType', 'ClusterStatus', 'ClusterVersion', 'NumberOfNodes', 'CreationDate'])
        for region in regions:
            client = session.client('redshift', region_name=region)
            try:
                clusters = client.describe_clusters().get('Clusters', [])
                if not clusters:
                    writer.writerow([region, 'No resource found', '', '', '', '', ''])
                for cluster in clusters:
                    writer.writerow([
                        region,
                        cluster['ClusterIdentifier'],
                        cluster['NodeType'],
                        cluster['ClusterStatus'],
                        cluster['ClusterVersion'],
                        cluster['NumberOfNodes'],
                        cluster['ClusterCreateTime'].strftime('%Y-%m-%d %H:%M:%S') if 'ClusterCreateTime' in cluster else 'N/A',
                    ])
            except ClientError as e:
                print(f"Error fetching Redshift clusters in {region}: {e}")

        writer.writerow([])

    # --------------------------SQS--------------------------
    elif resource_type == 'sqs':
        writer.writerow(['SQS Inventory'])
        writer.writerow(['Region', 'QueueUrl', 'Name', 'Type', 'ApproximateNumberOfMessages', 'CreatedDate'])
        for region in regions:
            client = session.client('sqs', region_name=region)
            try:
                queues = client.list_queues().get('QueueUrls', [])
                if not queues:
                    writer.writerow([region, 'No resource found', '', '', '', ''])
                for queue_url in queues:
                    attributes = client.get_queue_attributes(QueueUrl=queue_url, AttributeNames=['All']).get('Attributes', {})
                    queue_name = queue_url.split('/')[-1]
                    queue_type = attributes.get('QueueType', 'Standard')  # Default to 'Standard' if not specified
                    created_date = attributes.get('CreatedTimestamp', 'N/A')
                    if created_date != 'N/A':
                        created_date = datetime.utcfromtimestamp(float(created_date)).strftime('%Y-%m-%d %H:%M:%S')
                    writer.writerow([
                        region,
                        queue_url,
                        queue_name,
                        queue_type,
                        attributes.get('ApproximateNumberOfMessages', 'N/A'),
                        created_date,
                    ])
            except ClientError as e:
                print(f"Error fetching SQS queues in {region}: {e}")

        writer.writerow([])

    # -------------------------- Step Function--------------------------

    elif resource_type.lower() == 'stepfunctions':
        writer.writerow(['Step Functions Inventory'])
        writer.writerow(['Region', 'StateMachineArn', 'Name', 'Status', 'CreationDate'])
        for region in regions:
            client = session.client('stepfunctions', region_name=region)
            try:
                state_machines = client.list_state_machines().get('stateMachines', [])
                if not state_machines:
                    writer.writerow([region, 'No resource found', '', '', ''])
                for sm in state_machines:
                    writer.writerow([
                        region,
                        sm.get('stateMachineArn', 'N/A'),
                        sm.get('name', 'N/A'),
                        sm.get('status', 'N/A'),
                        sm.get('creationDate', datetime.min).strftime('%Y-%m-%d %H:%M:%S')
                        if sm.get('creationDate') else 'N/A',
                    ])
            except ClientError as e:
                print(f"Error fetching Step Functions in {region}: {e}")
                writer.writerow([region, 'Error fetching resources', '', '', ''])

        writer.writerow([])

    # -------------------------- Route 53--------------------------
    elif resource_type == 'route53':
        writer.writerow(['Route 53 Inventory'])
        writer.writerow(['HostedZoneId', 'Name', 'ResourceRecordSetCount', 'PrivateZone'])
        client = session.client('route53')
        try:
            zones = client.list_hosted_zones().get('HostedZones', [])
            if not zones:
                writer.writerow(['No resource found', '', '', ''])
            for zone in zones:
                writer.writerow([
                    zone['Id'],
                    zone['Name'],
                    zone['ResourceRecordSetCount'],
                    zone['Config']['PrivateZone'],
                ])
        except ClientError as e:
            print(f"Error fetching Route 53 zones: {e}")

        writer.writerow([])

    # --------------------------SNS--------------------------
    elif resource_type == 'sns':
        writer.writerow(['SNS Inventory'])
        writer.writerow(['Region', 'TopicArn', 'Name', 'Type', 'SubscriptionId', 'Endpoint', 'Status'])
        for region in regions:
            client = session.client('sns', region_name=region)
            try:
                topics = client.list_topics().get('Topics', [])
                if not topics:
                    writer.writerow([region, 'No resource found', '', '', '', '', ''])
                for topic in topics:
                    arn = topic['TopicArn']
                    attributes = client.get_topic_attributes(TopicArn=arn).get('Attributes', {})
                    subscriptions = client.list_subscriptions_by_topic(TopicArn=arn).get('Subscriptions', [])
                    if not subscriptions:
                        writer.writerow([region, arn, attributes.get('DisplayName', 'N/A'), attributes.get('TopicType', 'N/A'), 'No subscriptions found', '', ''])
                    for sub in subscriptions:
                        writer.writerow([
                            region,
                            arn,
                            attributes.get('DisplayName', 'N/A'),
                            attributes.get('TopicType', 'N/A'),
                            sub.get('SubscriptionArn', 'N/A'),
                            sub.get('Endpoint', 'N/A'),
                            sub.get('Protocol', 'N/A'),
                        ])
            except ClientError as e:
                print(f"Error fetching SNS topics in {region}: {e}")

        writer.writerow([])
        
    # -------------------------- Lambda Inventory --------------------------
    elif resource_type.lower() == 'lambda':
        writer.writerow(['Lambda Inventory'])
        writer.writerow(['Region', 'FunctionName', 'Runtime', 'Status', 'MemorySize', 'Timeout'])

        for region in regions:
            lambda_client = session.client('lambda', region_name=region)
            region_resources_found = False
            try:
                functions = lambda_client.list_functions().get('Functions', [])
                if not functions:
                    region_resources_found = True
                    writer.writerow([region, 'No resource found', '', '', '', ''])
                for function in functions:
                    region_resources_found = True
                    writer.writerow([region,
                                     function['FunctionName'],
                                     function['Runtime'],
                                     function['State'] if 'State' in function else 'N/A',
                                     function['MemorySize'],
                                     function['Timeout']])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', ''])

            except ClientError as e:
                print(f"Error fetching Lambda functions in {region}: {e}")

        writer.writerow([])

    # -------------------------- AWS Glue Inventory --------------------------
    elif resource_type.lower() == 'glue':
        writer.writerow(['AWS Glue Inventory'])
        writer.writerow(['Region', 'JobName', 'State', 'CreatedOn', 'LastModifiedOn', 'Role', 'Tags'])

        for region in regions:
            glue_client = session.client('glue', region_name=region)
            region_resources_found = False
            try:
                jobs = glue_client.get_jobs()['Jobs']
                if not jobs:
                    region_resources_found = True
                    writer.writerow([region, 'No resource found', '', '', '', '', ''])
                for job in jobs:
                    region_resources_found = True
                    writer.writerow([region,
                                     job['Name'],
                                     job['State'],
                                     job['CreatedOn'].strftime('%Y-%m-%d %H:%M:%S') if 'CreatedOn' in job else 'N/A',
                                     job['LastModifiedOn'].strftime('%Y-%m-%d %H:%M:%S') if 'LastModifiedOn' in job else 'N/A',
                                     job['Role'],
                                     get_tags(job.get('Tags', []))])
                if not region_resources_found:
                    writer.writerow([region, 'No resource found', '', '', '', '', ''])

            except ClientError as e:
                print(f"Error fetching Glue jobs in {region}: {e}")

        writer.writerow([])
        
    # -------------------------- EFS Inventory --------------------------
    elif resource_type == 'efs':
        # Elastic File System inventory
        writer.writerow(['EFS Inventory'])
        writer.writerow(['Region', 'FileSystemId', 'CreationTime', 'LifeCycleState', 'PerformanceMode', 'SizeInBytes'])
        for region in regions:
            client = session.client('efs', region_name=region)
            try:
                filesystems = client.describe_file_systems().get('FileSystems', [])
                if not filesystems:
                    writer.writerow([region, 'No resource found', '', '', '', ''])
                for fs in filesystems:
                    writer.writerow([
                        region,
                        fs['FileSystemId'],
                        fs['CreationTime'].strftime('%Y-%m-%d %H:%M:%S') if 'CreationTime' in fs else 'N/A',
                        fs['LifeCycleState'],
                        fs['PerformanceMode'],
                        fs['SizeInBytes']['Value'] if 'SizeInBytes' in fs else 'N/A',
                    ])
            except ClientError as e:
                print(f"Error fetching EFS in {region}: {e}")

        writer.writerow([])

    # -------------------------- Amplify Inventory --------------------------
    elif resource_type == 'amplify':
        # Amplify inventory
        writer.writerow(['Amplify Inventory'])
        writer.writerow(['Region', 'AppId', 'Name', 'CreationTime', 'Repository', 'Branch'])
        for region in regions:
            client = session.client('amplify', region_name=region)
            try:
                apps = client.list_apps().get('apps', [])
                if not apps:
                    writer.writerow([region, 'No resource found', '', '', '', ''])
                for app in apps:
                    writer.writerow([
                        region,
                        app['appId'],
                        app['name'],
                        app['createTime'].strftime('%Y-%m-%d %H:%M:%S') if 'createTime' in app else 'N/A',
                        app.get('repository', 'N/A'),
                        app.get('defaultDomain', 'N/A'),
                    ])
            except ClientError as e:
                print(f"Error fetching Amplify apps in {region}: {e}")

        writer.writerow([])
        
# -------------------------- S3 Inventory --------------------------
    elif resource_type == 's3':
        # CloudFront inventory
        writer.writerow(['S3 Inventory'])
        writer.writerow(['BucketName', 'CreationDate', 'Region', 'NumberOfObjects', 'Tags', 'LifecycleConfiguration'])

        client = session.client('s3')
        try:
            buckets = client.list_buckets().get('Buckets', [])
            for bucket in buckets:
                bucket_name = bucket['Name']
                creation_date = bucket['CreationDate']
                region = client.get_bucket_location(Bucket=bucket_name)['LocationConstraint'] or 'us-east-1'

                # Fetch tags
                tags = 'No Tags'
                try:
                    tags = get_tags(client.get_bucket_tagging(Bucket=bucket_name).get('TagSet', []))
                except ClientError:
                    pass

                # Fetch object count
                number_of_objects = client.list_objects_v2(Bucket=bucket_name).get('KeyCount', 0)

                # Fetch lifecycle configuration
                lifecycle_config = 'No Lifecycle Configuration'
                try:
                    lifecycle_rules = client.get_bucket_lifecycle_configuration(Bucket=bucket_name).get('Rules', [])
                    lifecycle_config = ', '.join([rule['ID'] for rule in lifecycle_rules if 'ID' in rule])
                except ClientError:
                    pass

                # Write row to CSV
                writer.writerow([bucket_name, creation_date, region, number_of_objects, tags, lifecycle_config])
        except ClientError as e:
            print(f"Error fetching S3 buckets: {e}")
        
        writer.writerow([])  # Blank row to separate sections

    # -------------------------- CloudFront Inventory --------------------------
    elif resource_type == 'cloudfront':
        # CloudFront inventory
        writer.writerow(['CloudFront Inventory'])
        writer.writerow(['DistributionId', 'DomainName', 'Status', 'Enabled', 'LastModifiedTime'])
        client = session.client('cloudfront')
        try:
            distributions = client.list_distributions().get('DistributionList', {}).get('Items', [])
            if not distributions:
                writer.writerow(['No resource found', '', '', '', ''])
            for dist in distributions:
                writer.writerow([
                    dist['Id'],
                    dist['DomainName'],
                    dist['Status'],
                    dist['Enabled'],
                    dist['LastModifiedTime'].strftime('%Y-%m-%d %H:%M:%S') if 'LastModifiedTime' in dist else 'N/A',
                ])
        except ClientError as e:
            print(f"Error fetching CloudFront distributions: {e}")

        writer.writerow([])


    # -------------------------- IAM User, Group, Policy, Role Inventory --------------------------
    elif resource_type.lower() in {'iam_user', 'iam_group', 'iam_policies', 'iam_role'}:
        iam_client = session.client('iam')
        resourcegroupstaggingapi_client = session.client('resourcegroupstaggingapi')

        if resource_type.lower() == 'iam_user':
            get_iam_user_inventory(iam_client, resourcegroupstaggingapi_client, writer)
        elif resource_type.lower() == "iam_group":
            get_iam_group_inventory(iam_client, resourcegroupstaggingapi_client, writer)
        elif resource_type.lower() == "iam_policies":
            get_iam_policy_inventory(iam_client, resourcegroupstaggingapi_client, writer)
        elif resource_type.lower() == "iam_role":
            get_iam_role_inventory(iam_client, resourcegroupstaggingapi_client, writer)
    else:
        print(f"Unsupported resource type: {resource_type}")
        return


def build_workbook_for_region(region_name, selected_resources, session, cost_period=None, interactive=False):
    if Workbook is None:
        raise RuntimeError("openpyxl is required for workbook generation.")

    region_label = region_name or "global"
    workbook = Workbook()
    sheet_created = False

    for resource_type in selected_resources:
        collector = ListWriter()
        try:
            generate_inventory_csv(
                [region_name],
                resource_type,
                collector,
                session,
                cost_period=cost_period,
                interactive=interactive,
            )
        except Exception as exc:
            collector.writerow([f"Error processing {resource_type}: {exc}"])

        if not collector.rows:
            continue

        rows_to_write = collector.rows
        table_blocks = extract_table_blocks(rows_to_write)

        if not table_blocks:
            error_messages = [
                [stringify_cell(row[0]).strip()]
                for row in rows_to_write
                if len(row) == 1 and "error" in stringify_cell(row[0]).lower()
            ]
            if error_messages:
                rows_to_write = [["Error Message"]] + error_messages
                table_blocks = [
                    {
                        "start": 0,
                        "end": len(rows_to_write) - 1,
                        "header_len": 1,
                        "max_cols": 1,
                    }
                ]
            else:
                continue

        sheet_name = sanitize_sheet_name(resource_type.replace('_', ' ').title())
        if not sheet_created:
            sheet = workbook.active
            sheet.title = sheet_name
            sheet_created = True
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        for row in rows_to_write:
            sheet.append(row)

        style_titles(sheet, rows_to_write)
        apply_table_formatting(sheet, rows_to_write, table_blocks)
        autofit_columns(sheet)

    if not sheet_created:
        return None

    return workbook, region_label


def generate_workbooks(
    session,
    regions,
    selected_resources,
    cost_period=None,
    interactive=False,
    max_workers=8,
    timestamp=None,
):
    if Workbook is None:
        raise RuntimeError("openpyxl is required for spreadsheet export.")
    if not regions:
        raise ValueError("At least one region must be provided.")
    if not selected_resources:
        raise ValueError("No resources selected for inventory generation.")

    used_timestamp = timestamp or datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    max_workers = max(1, min(max_workers, len(regions)))
    results = []

    def worker(region_name):
        workbook_tuple = build_workbook_for_region(
            region_name=region_name,
            selected_resources=selected_resources,
            session=session,
            cost_period=cost_period,
            interactive=interactive,
        )

        region_label = region_name or "global"
        if not workbook_tuple:
            return {"region": region_label, "filename": None, "content": None}

        workbook, region_label = workbook_tuple
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        safe_region = region_label.replace(" ", "_") or "global"
        filename = f"{safe_region}_inventory_{used_timestamp}.xlsx"
        return {"region": region_label, "filename": filename, "content": buffer.getvalue()}

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(worker, region) for region in regions]
        for future in as_completed(futures):
            results.append(future.result())

    return results, used_timestamp

def main():
    if Workbook is None:
        print("Required dependency 'openpyxl' is not installed. Install it with 'pip install openpyxl' and retry.")
        return

    args = parse_cli_args()

    try:
        access_key, secret_key, session_token, profile_name = resolve_credentials(args)

        session = create_boto3_session(
            access_key=access_key,
            secret_key=secret_key,
            session_token=session_token,
            profile_name=profile_name,
        )

        regions = parse_regions_input(args.regions)

        if not regions:
            regions_input = input("Enter regions separated by commas (leave empty for 'us-east-1'): ")
            regions = parse_regions_input(regions_input)

        if not regions:
            regions = ['us-east-1']
        
        selected_resources = parse_resource_selection(args.resources, RESOURCE_MAP)

        if not selected_resources:
            print("Available resources:")
            for number, resource in RESOURCE_MAP.items():
                print(f"{number}: {resource}")
            print("Enter 'all' to select all resources.")

            selection = input("Enter the numbers corresponding to the resources you want (e.g., 2,4) or 'all': ")
            selected_resources = parse_resource_selection(selection, RESOURCE_MAP)

        if not selected_resources:
            print("No valid resources selected. Exiting.")
            return

        cost_period = None
        if any(resource.lower() == 'cost' for resource in selected_resources):
            cost_period = resolve_cost_period_inputs(args.from_date, args.to_date)
            if not cost_period:
                start_input = input("Enter the start date for cost analysis (YYYY-MM-DD, DD-MM-YYYY, or 'last 30 days'): ").strip()
                end_input = input("Enter the end date for cost analysis (YYYY-MM-DD or DD-MM-YYYY, leave blank for today): ").strip()
                cost_period = resolve_cost_period(start_input, end_input)
            print(f"Using cost analysis window: {cost_period[0]} to {cost_period[1]}")

        print(f"[AWS Inventory] Starting inventory for regions: {', '.join(regions)}")
        print(f"[AWS Inventory] Selected resources: {', '.join(selected_resources)}")
        print("[AWS Inventory] Generating workbooks...")
        workbooks, timestamp = generate_workbooks(
            session=session,
            regions=regions,
            selected_resources=selected_resources,
            cost_period=cost_period,
            interactive=False,
        )

        output_dir = f"aws_inventory_{timestamp}"
        os.makedirs(output_dir, exist_ok=True)
        saved_any = False
        manifest_lines = ["Generated workbooks:"]

        for workbook_data in workbooks:
            region_label = workbook_data["region"]
            if not workbook_data["content"]:
                print(f"[AWS Inventory] No resources found for {region_label}; skipping workbook.")
                continue

            file_path = os.path.join(output_dir, workbook_data["filename"])
            with open(file_path, "wb") as file_obj:
                file_obj.write(workbook_data["content"])
            manifest_lines.append(f"- {workbook_data['filename']}")
            print(f"[AWS Inventory] Finished {region_label} -> {file_path}")
            saved_any = True

        if saved_any:
            readme_path = os.path.join(output_dir, "README.txt")
            with open(readme_path, "w", encoding="utf-8") as readme_file:
                readme_file.write("\n".join(manifest_lines))

            archive_path = shutil.make_archive(output_dir, "zip", output_dir)
            print(f"[AWS Inventory] Finished processing regions. Output folder: {output_dir}")
            print(f"[AWS Inventory] Archive created at {archive_path}")
        else:
            print("[AWS Inventory] No inventories were generated for the requested regions.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
